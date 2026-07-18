"""
panels/team_analytics.py — 4me team analytics, ported into CCP.

A focused head-to-head between POS Support and POS EFT: completed-ticket
volume per team and per person, a rising/falling trend, and a side-by-side
comparison. Auto-refreshes every 3 minutes in the background.

Structure:
  • If no saved credentials → inline connect form (token + account).
  • Once connected → period toggle (Today / Week / Month), two big team
    cards side by side, per-agent leaderboards under each team, and a
    trend line showing daily closes.

Credentials live in the shared 4me_pro base64 file so refresh is silent.
"""
from datetime import date, timedelta

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QPushButton,
    QLineEdit, QScrollArea, QFrame, QSizeGrip
)
from PyQt6.QtCore import (
    Qt, QTimer, QRectF, QPoint, QPropertyAnimation, QEasingCurve, pyqtProperty,
    pyqtSignal
)
from PyQt6.QtGui import QPainter, QColor, QPen, QPainterPath, QLinearGradient, QRadialGradient, QBrush

from src.widgets.xurrent_client import (
    AnalyticsWorker, vault_load, vault_save, vault_clear,
    today_str, week_start, month_start, TEAM_IDS,
)

FONT = "'Segoe UI Variable Text','Segoe UI Variable','Inter','Segoe UI',sans-serif"
MONO = "'JetBrains Mono','Cascadia Code','Consolas',monospace"

REFRESH_MS = 3 * 60 * 1000   # 3 minutes

# Distinct colours so the two teams read apart at a glance.
TEAM_COLORS = {
    "POS Support": "#38bdf8",   # blue
    "POS EFT":     "#a78bfa",   # violet
}

# Friendly labels for the 9 open RequestStatus values (confirmed directly
# against this 4me instance's own enum — see xurrent_client.build_open_count_query).
STATUS_LABELS = {
    "on_backlog":           "On backlog",
    "assigned":             "Assigned",
    "accepted":             "Accepted",
    "in_progress":          "In progress",
    "waiting_for":          "Waiting for",
    "waiting_for_customer": "Waiting for customer",
    "reservation_pending":  "Reservation pending",
    "workflow_pending":     "Workflow pending",
    "project_pending":      "Project pending",
}
# Display order — roughly the lifecycle order a call moves through.
STATUS_ORDER = list(STATUS_LABELS.keys())


class _ClickableBox(QWidget):
    """A QWidget that behaves like a button — used for the open-call badges
    so clicking one toggles its per-status breakdown."""
    clicked = pyqtSignal()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit()
        super().mousePressEvent(event)


# ── tiny sparkline / trend line ───────────────────────────────────────────────
class TrendLine(QWidget):
    """
    Clean daily-close trend chart for the two teams. Designed for a light
    background: soft gridlines, a Y-axis scale, day labels along the bottom,
    a legend, a value dot on each series' latest point, and light gradient
    fills that don't muddy each other.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(200)
        self._series = {}          # team -> list[(day, count)]

    def set_data(self, series: dict, accent: str = None):
        self._series = series
        self.update()

    def _nice_step(self, vmax):
        import math
        if vmax <= 5:
            return 1
        raw = vmax / 4.0
        mag = 10 ** int(math.floor(math.log10(raw)))
        for m in (1, 2, 5, 10):
            if raw <= m * mag:
                return int(m * mag)
        return int(10 * mag)

    def paintEvent(self, _):
        from PyQt6.QtGui import QFont
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w, h = self.width(), self.height()
        ml, mr, mt, mb = 42, 16, 20, 30
        plot_w = w - ml - mr
        plot_h = h - mt - mb

        all_vals = [c for pts in self._series.values() for _, c in pts]
        if not all_vals or plot_w <= 0:
            p.setPen(QColor("#94a3b8"))
            p.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, "No trend data yet")
            return
        vmax = max(all_vals)
        step = self._nice_step(vmax)
        top = step * (int(vmax / step) + 1) if vmax % step else vmax
        top = max(top, step)

        small = QFont(FONT.split(',')[0].strip("'"), 8)
        p.setFont(small)

        # horizontal gridlines + Y labels
        n_lines = int(top / step)
        for i in range(n_lines + 1):
            val = step * i
            y = mt + plot_h - (plot_h * (val / top))
            p.setPen(QPen(QColor(0, 0, 0, 20), 1))
            p.drawLine(ml, int(y), ml + plot_w, int(y))
            p.setPen(QColor("#64748b"))
            p.drawText(0, int(y) - 7, ml - 6, 14,
                       Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter,
                       str(int(val)))

        # day axis
        days = []
        for pts in self._series.values():
            if len(pts) > len(days):
                days = [d for d, _ in pts]
        n = len(days)
        if n > 1:
            label_every = max(1, n // 8)
            p.setPen(QColor("#64748b"))
            for i, d in enumerate(days):
                if i % label_every == 0 or i == n - 1:
                    x = ml + plot_w * (i / (n - 1))
                    txt = d[5:] if len(d) >= 10 else d
                    p.drawText(int(x) - 18, mt + plot_h + 6, 36, 14,
                               Qt.AlignmentFlag.AlignCenter, txt)

        # each team line + soft fill
        for team, pts in self._series.items():
            if len(pts) < 2:
                continue
            color = QColor(TEAM_COLORS.get(team, "#38bdf8"))
            m = len(pts)
            line = QPainterPath(); fill = QPainterPath()
            for i, (_, c) in enumerate(pts):
                x = ml + plot_w * (i / (m - 1))
                y = mt + plot_h - (plot_h * (c / top))
                if i == 0:
                    line.moveTo(x, y); fill.moveTo(x, mt + plot_h); fill.lineTo(x, y)
                else:
                    line.lineTo(x, y); fill.lineTo(x, y)
            fill.lineTo(ml + plot_w, mt + plot_h); fill.closeSubpath()
            grad = QLinearGradient(0, mt, 0, mt + plot_h)
            fc = QColor(color); fc.setAlpha(38)
            fc2 = QColor(color); fc2.setAlpha(0)
            grad.setColorAt(0, fc); grad.setColorAt(1, fc2)
            p.setPen(Qt.PenStyle.NoPen); p.setBrush(grad); p.drawPath(fill)
            p.setPen(QPen(color, 2.4)); p.setBrush(Qt.BrushStyle.NoBrush); p.drawPath(line)
            _, lc = pts[-1]
            lx = ml + plot_w; ly = mt + plot_h - (plot_h * (lc / top))
            p.setBrush(color); p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(int(lx) - 4, int(ly) - 4, 8, 8)

        # legend (top-right)
        p.setFont(small)
        lx = ml + plot_w
        for team in reversed(list(self._series.keys())):
            color = QColor(TEAM_COLORS.get(team, "#38bdf8"))
            tw = p.fontMetrics().horizontalAdvance(team)
            lx -= tw + 20
            p.setBrush(color); p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(int(lx), mt - 12, 8, 8)
            p.setPen(QColor("#475569"))
            p.drawText(int(lx) + 12, mt - 14, tw + 6, 14,
                       Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, team)



# ── team head-to-head card ────────────────────────────────────────────────────
class TeamCard(QWidget):
    """Big number for one team + its per-agent leaderboard."""
    def __init__(self, team: str, parent=None):
        super().__init__(parent)
        self.team = team
        self.color = TEAM_COLORS.get(team, "#38bdf8")
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setStyleSheet(
            f"background:qlineargradient(x1:0,y1:0,x2:0,y2:1,"
            f"stop:0 rgba(255,255,255,0.03),stop:0.12 #0d1520,stop:1 #0d1520);"
            f"border:1px solid #172338;border-top:1px solid {self.color}66;"
            f"border-radius:12px;")

        self.setMinimumHeight(300)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(18, 16, 18, 16)
        lay.setSpacing(4)

        head = QLabel(team.upper())
        head.setStyleSheet(
            f"background:transparent;color:{self.color};font-size:11px;"
            f"font-weight:700;letter-spacing:2px;border:none;")
        lay.addWidget(head)

        self._big = QLabel("0")
        self._big.setStyleSheet(
            f"background:transparent;color:#e8eef5;font-size:44px;"
            f"font-weight:900;border:none;font-family:{MONO};")
        lay.addWidget(self._big)
        self._display_val = 0
        self._count_anim = None

        self._sub = QLabel("tickets closed")
        self._sub.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;border:none;")
        lay.addWidget(self._sub)

        self._delta = QLabel("")
        self._delta.setStyleSheet(
            "background:transparent;font-size:10px;font-weight:600;border:none;")
        lay.addWidget(self._delta)

        # metric chips: active agents · median speed · per-agent throughput
        self._chips = QLabel("")
        self._chips.setStyleSheet(
            f"background:transparent;color:#8ba0b6;font-size:10px;border:none;"
            f"font-family:{MONO};")
        lay.addWidget(self._chips)

        lay.addSpacing(6)
        agents_lbl = QLabel("BY AGENT")
        agents_lbl.setStyleSheet(
            "background:transparent;color:#6b83a0;font-size:10px;"
            "font-weight:700;letter-spacing:1.5px;border:none;")
        lay.addWidget(agents_lbl)

        self._agents_box = QVBoxLayout()
        self._agents_box.setSpacing(3)
        lay.addLayout(self._agents_box)
        lay.addStretch()

    @pyqtProperty(int)
    def countVal(self):
        return self._display_val

    @countVal.setter
    def countVal(self, v):
        self._display_val = v
        self._big.setText(str(v))

    def _animate_count(self, target):
        try:
            self._count_anim = QPropertyAnimation(self, b"countVal")
            self._count_anim.setDuration(700)
            self._count_anim.setStartValue(self._display_val)
            self._count_anim.setEndValue(int(target))
            self._count_anim.setEasingCurve(QEasingCurve.Type.OutCubic)
            self._count_anim.start()
        except Exception:
            self._display_val = int(target)
            self._big.setText(str(int(target)))

    def set_value(self, total: int, agents: dict, delta_txt: str = "",
                  delta_up=None, metrics: dict = None):
        self._animate_count(total)
        self._sub.setText("ticket closed" if total == 1 else "tickets closed")

        if delta_txt:
            col = "#00e87a" if delta_up else ("#f87171" if delta_up is False else "#6d84a0")
            arrow = "▲ " if delta_up else ("▼ " if delta_up is False else "")
            self._delta.setText(f"{arrow}{delta_txt}")
            self._delta.setStyleSheet(
                f"background:transparent;color:{col};font-size:10px;"
                f"font-weight:600;border:none;")
        else:
            self._delta.setText("")

        # metric chips
        m = metrics or {}
        active = m.get("active", 0)
        per    = m.get("per_agent", 0.0)
        med    = m.get("med_hours")
        def _fmt_h(h):
            if h is None: return "—"
            if h < 1:  return f"{h*60:.0f}m"
            if h < 48: return f"{h:.1f}h"
            return f"{h/24:.1f}d"
        self._chips.setText(
            f"{active} active   ·   {per:.1f}/agent   ·   ⌀ {_fmt_h(med)}")

        # rebuild agent rows
        while self._agents_box.count():
            it = self._agents_box.takeAt(0)
            if it.widget():
                it.widget().setParent(None)

        if not agents:
            empty = QLabel("no closes in period")
            empty.setStyleSheet(
                "background:transparent;color:#6b83a0;font-size:10px;"
                "font-style:italic;border:none;")
            self._agents_box.addWidget(empty)
            return

        ranked = sorted(agents.items(), key=lambda kv: kv[1], reverse=True)
        top = ranked[0][1] if ranked else 1
        for i, (name, count) in enumerate(ranked):
            self._agents_box.addWidget(self._agent_row(name, count, top, i + 1))

    def _agent_row(self, name, count, top, rank):
        row = QWidget()
        row.setStyleSheet(
            "QWidget{background:transparent;border-radius:8px;}"
            "QWidget:hover{background:rgba(255,255,255,0.03);}")
        rl = QHBoxLayout(row)
        rl.setContentsMargins(4, 2, 4, 2)
        rl.setSpacing(8)

        # rank badge — medal tint for top 3
        medal = {1: "#ffd54a", 2: "#c0cbd8", 3: "#e0955a"}.get(rank)
        rk = QLabel(str(rank))
        rk.setFixedWidth(18)
        rk.setAlignment(Qt.AlignmentFlag.AlignCenter)
        if medal:
            rk.setStyleSheet(
                f"background:transparent;color:{medal};font-size:10px;"
                f"font-weight:800;border:none;font-family:{MONO};")
        else:
            rk.setStyleSheet(
                "background:transparent;color:#6b83a0;font-size:10px;"
                "border:none;font-family:'JetBrains Mono',monospace;")
        rl.addWidget(rk)

        nm = QLabel(name)
        nm.setStyleSheet(
            f"background:transparent;color:{'#e8eef5' if rank <= 3 else '#9fb0c4'};"
            f"font-size:11px;font-weight:{'600' if rank <= 3 else '400'};border:none;")
        rl.addWidget(nm, 1)

        bar = _MiniBar(count / top if top else 0, self.color)
        rl.addWidget(bar)

        val = QLabel(str(count))
        val.setFixedWidth(30)
        val.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        val.setStyleSheet(
            f"background:transparent;color:{self.color};font-size:12px;"
            f"font-weight:800;border:none;font-family:{MONO};")
        rl.addWidget(val)
        return row


class _MiniBar(QWidget):
    def __init__(self, frac, color, parent=None):
        super().__init__(parent)
        self._frac = max(0.0, min(1.0, frac))
        self._color = QColor(color)
        self.setFixedSize(70, 6)

    def paintEvent(self, _):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w, h = self.width(), self.height()
        p.setBrush(QColor(255, 255, 255, 14))
        p.setPen(Qt.PenStyle.NoPen)
        p.drawRoundedRect(0, 0, w, h, 3, 3)
        fw = int(w * self._frac)
        if fw > 0:
            grad = QLinearGradient(0, 0, fw, 0)
            c1 = QColor(self._color); c1.setAlpha(150)
            c2 = QColor(self._color)
            grad.setColorAt(0, c1); grad.setColorAt(1, c2)
            p.setBrush(QBrush(grad))
            p.drawRoundedRect(0, 0, max(fw, 4), h, 3, 3)


# ── main panel ────────────────────────────────────────────────────────────────
class Scoreboard(QWidget):
    """
    Head-to-head verdict panel. Each metric row reads left-to-right in plain
    English: METRIC | what it means | Team A value | who wins | Team B value.
    The winning team's chip lights up; a summary banner declares the overall
    leader by weighted metric wins.

    Metrics in priority order (weight): Speed 4, Throughput 3,
    Consistency 2, Balance 1.
    """
    # (key, label, hint, higher_is_better, kind, weight, tooltip)
    METRICS = [
        ("med_hours",  "Speed",       "median time to close",      False, "hours", 4,
         "SPEED — median time to close a ticket.\n\n"
         "Uses 4me's resolutionDuration (created → completed). The median is "
         "used instead of the average so a few very slow tickets don't skew it.\n\n"
         "Lower is better. The clearest single sign of an effective team."),
        ("per_agent",  "Throughput",  "closed per active person",  True,  "per",   3,
         "THROUGHPUT — tickets closed per active person.\n\n"
         "Total closed ÷ number of people who closed at least one ticket. This "
         "corrects for team size: a bigger team closing more in total isn't "
         "necessarily more productive per head.\n\n"
         "Higher is better."),
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:0,y2:1,"
            "stop:0 rgba(255,255,255,0.035),stop:0.14 rgba(15,23,36,0.9),"
            "stop:1 rgba(15,23,36,0.9));"
            "border:1px solid #1c2b42;border-top:1px solid rgba(255,255,255,0.10);"
            "border-radius:12px;")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 16, 20, 18)
        lay.setSpacing(0)

        # verdict banner
        self._verdict = QLabel("Waiting for data…")
        self._verdict.setStyleSheet(
            f"background:transparent;color:#6d84a0;font-size:15px;"
            f"font-weight:800;border:none;font-family:{FONT};")
        lay.addWidget(self._verdict)

        self._verdict_sub = QLabel("")
        self._verdict_sub.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;border:none;")
        lay.addWidget(self._verdict_sub)
        lay.addSpacing(14)

        # column headers (team names) — set once teams known
        self._hdr = QHBoxLayout()
        self._hdr.setContentsMargins(0, 0, 0, 0)
        lay.addLayout(self._hdr)
        lay.addSpacing(4)

        self._rows_box = QVBoxLayout()
        self._rows_box.setSpacing(0)
        lay.addLayout(self._rows_box)

    def _fmt(self, val, kind):
        if val is None:
            return "—"
        if kind == "hours":
            if val < 1:   return f"{val*60:.0f} min"
            if val < 48:  return f"{val:.1f} hrs"
            return f"{val/24:.1f} days"
        if kind == "per":
            return f"{val:.1f}"
        return f"{val:.0f}%"

    def _build_header(self, ta, tb):
        while self._hdr.count():
            it = self._hdr.takeAt(0)
            if it.widget():
                it.widget().setParent(None)
        _spacer_css = "background:transparent;border:none;"
        spacer = QLabel("")
        spacer.setFixedWidth(230)
        spacer.setStyleSheet(_spacer_css)
        self._hdr.addWidget(spacer)
        a = QLabel(ta.upper())
        a.setAlignment(Qt.AlignmentFlag.AlignRight)
        a.setStyleSheet(
            f"background:transparent;color:{TEAM_COLORS.get(ta)};font-size:10px;"
            f"font-weight:800;letter-spacing:1px;border:none;")
        self._hdr.addWidget(a, 1)
        mid = QLabel("")
        mid.setFixedWidth(70)
        mid.setStyleSheet(_spacer_css)
        self._hdr.addWidget(mid)
        b = QLabel(tb.upper())
        b.setAlignment(Qt.AlignmentFlag.AlignLeft)
        b.setStyleSheet(
            f"background:transparent;color:{TEAM_COLORS.get(tb)};font-size:10px;"
            f"font-weight:800;letter-spacing:1px;border:none;")
        self._hdr.addWidget(b, 1)

    def update_metrics(self, metrics: dict, teams: list):
        while self._rows_box.count():
            it = self._rows_box.takeAt(0)
            if it.widget():
                it.widget().setParent(None)
        if len(teams) < 2:
            return
        ta, tb = teams[0], teams[1]
        self._build_header(ta, tb)
        ma, mb = metrics.get(ta, {}), metrics.get(tb, {})

        wins = {ta: 0.0, tb: 0.0}
        won_labels = {ta: [], tb: []}
        for key, label, hint, higher, kind, weight, tip in self.METRICS:
            va, vb = ma.get(key), mb.get(key)
            winner = None
            if va is not None and vb is not None and va != vb:
                a_better = (va > vb) if higher else (va < vb)
                winner = ta if a_better else tb
                wins[winner] += weight
                won_labels[winner].append(label)
            self._rows_box.addWidget(
                self._metric_row(label, hint, ta, tb, va, vb, kind, winner, tip))

        # verdict
        if wins[ta] == wins[tb]:
            self._verdict.setText("Evenly matched this period")
            self._verdict.setStyleSheet(
                f"background:transparent;color:#8ba0b6;font-size:15px;"
                f"font-weight:800;border:none;font-family:{FONT};")
            self._verdict_sub.setText("Neither team leads on the weighted metrics.")
            return None
        else:
            champ = ta if wins[ta] > wins[tb] else tb
            col = TEAM_COLORS.get(champ, "#38bdf8")
            self._verdict.setText(f"{champ} is ahead this period")
            self._verdict.setStyleSheet(
                f"background:transparent;color:{col};font-size:15px;"
                f"font-weight:800;border:none;font-family:{FONT};")
            wl = won_labels[champ]
            self._verdict_sub.setText(
                "Leads on " + ", ".join(wl) + "." if wl else "")
            return champ

    def _metric_row(self, label, hint, ta, tb, va, vb, kind, winner, tip=""):
        row = QWidget()
        row.setStyleSheet(
            "QWidget{background:transparent;}"
            "QWidget:hover{background:rgba(255,255,255,0.02);border-radius:8px;}")
        if tip:
            row.setToolTip(tip)
            row.setCursor(Qt.CursorShape.WhatsThisCursor)
        rl = QHBoxLayout(row)
        rl.setContentsMargins(6, 7, 6, 7)
        rl.setSpacing(0)

        # label + hint block (fixed width)
        lblbox = QVBoxLayout()
        lblbox.setSpacing(0)
        name_row = QHBoxLayout(); name_row.setSpacing(5); name_row.setContentsMargins(0,0,0,0)
        name = QLabel(label)
        name.setStyleSheet(
            "background:transparent;color:#e2e8f0;font-size:12px;"
            "font-weight:700;border:none;")
        name_row.addWidget(name)
        info = QLabel("ⓘ")
        info.setStyleSheet(
            "background:transparent;color:#6b83a0;font-size:10px;border:none;")
        if tip:
            info.setToolTip(tip)
        name_row.addWidget(info)
        name_row.addStretch()
        lblbox.addLayout(name_row)
        hnt = QLabel(hint)
        hnt.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;border:none;")
        lblbox.addWidget(hnt)
        lw = QWidget(); lw.setFixedWidth(224); lw.setLayout(lblbox)
        lw.setStyleSheet("background:transparent;")
        rl.addWidget(lw)

        a_win = (winner == ta)
        b_win = (winner == tb)

        va_lbl = QLabel(self._fmt(va, kind))
        va_lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        va_lbl.setStyleSheet(
            f"background:transparent;color:{TEAM_COLORS.get(ta) if a_win else '#7d8fa3'};"
            f"font-size:14px;font-weight:{'800' if a_win else '600'};"
            f"border:none;font-family:{MONO};")
        rl.addWidget(va_lbl, 1)

        # centre win indicator
        mid = QLabel("◀" if a_win else ("▶" if b_win else "="))
        mid.setFixedWidth(70)
        mid.setAlignment(Qt.AlignmentFlag.AlignCenter)
        mid_col = (TEAM_COLORS.get(ta) if a_win else
                   TEAM_COLORS.get(tb) if b_win else "#6b83a0")
        mid.setStyleSheet(
            f"background:transparent;color:{mid_col};font-size:12px;"
            f"font-weight:800;border:none;")
        rl.addWidget(mid)

        vb_lbl = QLabel(self._fmt(vb, kind))
        vb_lbl.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        vb_lbl.setStyleSheet(
            f"background:transparent;color:{TEAM_COLORS.get(tb) if b_win else '#7d8fa3'};"
            f"font-size:14px;font-weight:{'800' if b_win else '600'};"
            f"border:none;font-family:{MONO};")
        rl.addWidget(vb_lbl, 1)
        return row


class PeriodGlance(QWidget):
    """
    A compact strip showing both teams' closed totals across all three
    periods at once — Today / Week / Month — so you never have to click
    the toggle just to compare. Cells fill in as each period is fetched.
    """
    PERIODS = [("today", "TODAY"), ("week", "THIS WEEK"), ("month", "THIS MONTH")]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setStyleSheet(
            "background:rgba(11,17,27,0.7);border:1px solid #172338;"
            "border-radius:12px;")
        self._cells = {}
        lay = QHBoxLayout(self)
        lay.setContentsMargins(6, 10, 6, 10)
        lay.setSpacing(0)

        teams = list(TEAM_IDS.keys())
        for i, (key, label) in enumerate(self.PERIODS):
            cell = QWidget()
            cell.setStyleSheet("background:transparent;")
            cl = QVBoxLayout(cell)
            cl.setContentsMargins(14, 2, 14, 2)
            cl.setSpacing(3)

            head = QLabel(label)
            head.setAlignment(Qt.AlignmentFlag.AlignCenter)
            head.setStyleSheet(
                "background:transparent;color:#6d84a0;font-size:10px;"
                "font-weight:700;letter-spacing:1.5px;border:none;")
            cl.addWidget(head)

            row = QHBoxLayout()
            row.setSpacing(10)
            row.addStretch()
            t_lbls = {}
            for ti, team in enumerate(teams):
                v = QLabel("·")
                v.setStyleSheet(
                    f"background:transparent;color:{TEAM_COLORS.get(team)};"
                    f"font-size:20px;font-weight:800;border:none;font-family:{MONO};")
                t_lbls[team] = v
                row.addWidget(v)
                if ti == 0:
                    sep = QLabel("vs")
                    sep.setStyleSheet(
                        "background:transparent;color:#6b83a0;font-size:10px;border:none;")
                    row.addWidget(sep)
            row.addStretch()
            cl.addLayout(row)
            self._cells[key] = t_lbls
            lay.addWidget(cell, 1)

            if i < len(self.PERIODS) - 1:
                div = QFrame()
                div.setFrameShape(QFrame.Shape.VLine)
                div.setStyleSheet("background:#172338;border:none;max-width:1px;")
                lay.addWidget(div)

    def set_period(self, period_key, totals: dict):
        cells = self._cells.get(period_key)
        if not cells:
            return
        for team, lbl in cells.items():
            lbl.setText(str(totals.get(team, 0)))


class _ShareBar(QWidget):
    """Proportion bar: this team's share of total reassignments vs the other."""
    def __init__(self, frac, color, parent=None):
        super().__init__(parent)
        self._frac = max(0.0, min(1.0, frac))
        self._color = QColor(color)
        self.setFixedHeight(6)
        self.setMinimumWidth(80)

    def paintEvent(self, _):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w, h = self.width(), self.height()
        p.setPen(Qt.PenStyle.NoPen)
        p.setBrush(QColor(255, 255, 255, 16))
        p.drawRoundedRect(0, 0, w, h, 3, 3)
        fw = int(w * self._frac)
        if fw > 0:
            grad = QLinearGradient(0, 0, fw, 0)
            c1 = QColor(self._color)
            c2 = QColor(self._color); c2.setAlpha(160)
            grad.setColorAt(0, c2); grad.setColorAt(1, c1)
            p.setBrush(grad)
            p.drawRoundedRect(0, 0, max(fw, 5), h, 3, 3)


class HandoffPanel(QWidget):
    """
    Cross-team reassignments — which team absorbs work handed over from the
    other. Two team cards side by side, each with its absorbed count, a share
    bar showing its slice of all reassignments, and the raw move count. A
    verdict line reads the result in plain English.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:0,y2:1,"
            "stop:0 rgba(255,255,255,0.025),stop:0.14 rgba(13,21,32,0.7),"
            "stop:1 rgba(13,21,32,0.7));"
            "border:1px solid #172338;border-top:1px solid rgba(255,255,255,0.08);"
            "border-radius:12px;")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 14, 20, 16)
        lay.setSpacing(4)

        title = QLabel("CROSS-TEAM HANDOFFS")
        title.setStyleSheet(
            "background:transparent;color:#7d8fa3;font-size:10px;"
            "font-weight:800;letter-spacing:2px;border:none;")
        lay.addWidget(title)
        sub = QLabel("tickets each team absorbed after being reassigned from the other")
        sub.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;border:none;")
        lay.addWidget(sub)
        lay.addSpacing(12)

        self._cards_row = QHBoxLayout()
        self._cards_row.setSpacing(12)
        lay.addLayout(self._cards_row)

        self._verdict = QLabel("Loading handoff data…")
        self._verdict.setWordWrap(True)
        self._verdict.setStyleSheet(
            "background:transparent;color:#6b83a0;font-size:11px;"
            "font-style:italic;border:none;")
        lay.addSpacing(12)
        lay.addWidget(self._verdict)

        self._note = QLabel("")
        self._note.setWordWrap(True)
        self._note.setStyleSheet(
            "background:transparent;color:#6b83a0;font-size:9px;border:none;")
        lay.addWidget(self._note)

    def _clear_cards(self):
        while self._cards_row.count():
            it = self._cards_row.takeAt(0)
            if it.widget():
                it.widget().setParent(None)

    def set_metrics(self, metrics: dict, teams: list):
        self._clear_cards()
        if len(teams) < 2:
            return
        a, b = teams[0], teams[1]
        ma, mb = metrics.get(a, {}), metrics.get(b, {})
        a_in = ma.get("reassigned_in", 0)
        b_in = mb.get("reassigned_in", 0)
        a_moves = ma.get("reassign_moves", 0)
        b_moves = mb.get("reassign_moves", 0)
        total = (a_in + b_in) or 1

        self._cards_row.addWidget(
            self._team_card(a, a_in, a_moves, a_in / total, TEAM_COLORS.get(a)), 1)
        self._cards_row.addWidget(
            self._team_card(b, b_in, b_moves, b_in / total, TEAM_COLORS.get(b)), 1)

        if a_in == b_in:
            self._verdict.setText(f"Balanced — both teams absorbed {a_in} reassigned tickets.")
            self._verdict.setStyleSheet(
                "background:transparent;color:#8ba0b6;font-size:12px;"
                "font-weight:700;border:none;")
        else:
            taker = a if a_in > b_in else b
            other = b if a_in > b_in else a
            diff  = abs(a_in - b_in)
            pct = round(100 * max(a_in, b_in) / total)
            col = TEAM_COLORS.get(taker, "#38bdf8")
            self._verdict.setText(
                f"{taker} absorbs {pct}% of all reassigned work "
                f"— {diff} more tickets than {other}.")
            self._verdict.setStyleSheet(
                f"background:transparent;color:{col};font-size:12px;"
                f"font-weight:700;border:none;")

        self._note.setText(
            "Based on 4me assignmentCount > 1 (ticket changed teams before completing). "
            "The API exposes how many reassignments a ticket had, not the direction, "
            "so this shows which team received reassigned work — not who sent it.")

    def _team_card(self, team, count, moves, share, color):
        card = QWidget()
        card.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        card.setStyleSheet(
            f"background:rgba(255,255,255,0.02);border:1px solid #172338;"
            f"border-left:3px solid {color};border-radius:12px;")
        cl = QVBoxLayout(card)
        cl.setContentsMargins(16, 12, 16, 12)
        cl.setSpacing(6)

        name = QLabel(team.upper())
        name.setStyleSheet(
            f"background:transparent;color:{color};font-size:10px;"
            f"font-weight:800;letter-spacing:1px;border:none;")
        cl.addWidget(name)

        big = QLabel(str(count))
        big.setStyleSheet(
            f"background:transparent;color:#e8eef5;font-size:30px;"
            f"font-weight:800;border:none;font-family:{MONO};")
        cl.addWidget(big)

        cap = QLabel("reassigned tickets absorbed")
        cap.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;border:none;")
        cl.addWidget(cap)

        cl.addSpacing(4)
        bar = _ShareBar(share, color)
        cl.addWidget(bar)

        foot = QHBoxLayout(); foot.setContentsMargins(0, 0, 0, 0)
        pct = QLabel(f"{round(share*100)}% of all handoffs")
        pct.setStyleSheet(
            "background:transparent;color:#7d8fa3;font-size:10px;border:none;")
        foot.addWidget(pct)
        foot.addStretch()
        mv = QLabel(f"{moves} moves")
        mv.setStyleSheet(
            f"background:transparent;color:{color};font-size:10px;"
            f"font-weight:700;border:none;font-family:{MONO};")
        foot.addWidget(mv)
        cl.addLayout(foot)
        return card





class TeamAnalyticsWindow(QWidget):
    """
    Independent, draggable popout window (like SettingsWindow). Frameless
    Tool window, slides in from the right of the main window. It has a
    fixed minimum size that keeps every data element visible, and cannot
    be maximized/fullscreened — only resized larger, never smaller than
    the point where the two team cards + leaderboards + trend still fit.
    """
    MIN_W = 900
    MIN_H = 680

    def __init__(self, app_win):
        super().__init__(None)
        self._app = app_win
        self.app = app_win               # alias used by dashboard code
        self._palette = {}
        self._token = ""
        self._account = ""
        self._period = "today"           # today | week | month
        self._worker = None
        self._last_totals = {}
        # cached results per period so switching tabs is instant
        self._period_cache = {}

        # Native OS window: real title bar with minimize / maximize / close,
        # fully resizable, fullscreen-capable. No custom chrome.
        self.setWindowTitle("Team Analytics — POS Support vs POS EFT")
        self.setWindowFlags(Qt.WindowType.Window)
        from src.widgets.app_icon import apply_window_icon
        apply_window_icon(self)
        self.setMinimumSize(self.MIN_W, self.MIN_H)
        self.resize(1040, 720)

        self.setStyleSheet(
            "QWidget{background:transparent;color:#d4dfe9;"
            f"font-family:{FONT};font-size:12px;border:none;}}"
            "QScrollArea{background:transparent;border:none;}"
            "QScrollArea>QWidget{background:transparent;}"
            "QScrollBar:vertical{background:transparent;width:8px;margin:2px;}"
            "QScrollBar::handle:vertical{background:#1a2840;border-radius:4px;min-height:30px;}"
            "QScrollBar::handle:vertical:hover{background:#25406a;}"
            "QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical{height:0;}"
            "QToolTip{background:#0d1520;color:#d4dfe9;border:1px solid #2a3f5c;"
            "border-radius:8px;padding:8px 10px;font-size:11px;}"
        )

        # Everything scrolls, so nothing clips at small sizes.
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        body = QWidget()
        self._root = QVBoxLayout(body)
        self._root.setContentsMargins(24, 20, 24, 20)
        self._root.setSpacing(16)
        scroll.setWidget(body)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)
        outer.addWidget(scroll)

        self._timer = QTimer(self)
        self._timer.setInterval(REFRESH_MS)
        self._timer.timeout.connect(self._fetch)

        # Animated aurora background — subtle drifting blobs behind content.
        self._aurora_phase = 0.0
        self._aurora_timer = QTimer(self)
        self._aurora_timer.setInterval(50)   # 20fps is plenty for slow drift
        self._aurora_timer.timeout.connect(self._tick_aurora)
        self._aurora_timer.start()

        self._build_ui()

    def _tick_aurora(self):
        self._aurora_phase += 0.006
        self.update()

    def paintEvent(self, event):
        import math
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w, h = self.width(), self.height()

        # base deep gradient
        base = QLinearGradient(0, 0, 0, h)
        base.setColorAt(0, QColor("#0a0f1a"))
        base.setColorAt(0.5, QColor("#070b12"))
        base.setColorAt(1, QColor("#05080f"))
        p.fillRect(0, 0, w, h, QBrush(base))

        # drifting aurora blobs (team colours, very low alpha)
        ph = self._aurora_phase
        blobs = [
            (0.25 + 0.08 * math.sin(ph),       0.2 + 0.06 * math.cos(ph * 0.8),  "#38bdf8", 0.10),
            (0.75 + 0.07 * math.cos(ph * 0.9), 0.3 + 0.05 * math.sin(ph * 1.1), "#a78bfa", 0.10),
            (0.5  + 0.10 * math.sin(ph * 0.7), 0.7 + 0.05 * math.cos(ph),        "#00e87a", 0.05),
        ]
        for fx, fy, col, alpha in blobs:
            cx, cy = fx * w, fy * h
            r = min(w, h) * 0.55
            grad = QRadialGradient(cx, cy, r)
            c = QColor(col)
            c.setAlphaF(alpha)
            c2 = QColor(col); c2.setAlpha(0)
            grad.setColorAt(0, c)
            grad.setColorAt(1, c2)
            p.setBrush(QBrush(grad))
            p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(int(cx - r), int(cy - r), int(r * 2), int(r * 2))

    def open_window(self):
        """Show + focus the window (called by the placeholder panel)."""
        self.show()
        self.setWindowState(Qt.WindowState.WindowNoState)
        self.raise_()
        self.activateWindow()
        if hasattr(self, "_aurora_timer"):
            self._aurora_timer.start()

    def closeEvent(self, e):
        # Hide instead of destroy so state/cache persist between opens.
        self._timer.stop()
        if hasattr(self, "_aurora_timer"):
            self._aurora_timer.stop()   # don't animate while hidden
        e.ignore()
        self.hide()

    # ── UI construction ──────────────────────────────────────────
    def _clear_root(self):
        while self._root.count():
            it = self._root.takeAt(0)
            if it.widget():
                it.widget().setParent(None)
            elif it.layout():
                self._drop_layout(it.layout())

    def _drop_layout(self, lay):
        while lay.count():
            it = lay.takeAt(0)
            if it.widget():
                it.widget().setParent(None)
            elif it.layout():
                self._drop_layout(it.layout())

    def _build_ui(self):
        self._token, self._account = vault_load()
        self._clear_root()
        if self._token and self._account:
            self._build_dashboard()
            QTimer.singleShot(100, self._fetch)
            self._timer.start()
        else:
            self._timer.stop()
            self._build_login()

    def _build_login(self):
        self._timer.stop()
        wrap = QWidget()
        wrap.setStyleSheet("background:transparent;")
        wl = QVBoxLayout(wrap)
        wl.setContentsMargins(0, 0, 0, 0)
        wl.addStretch()

        card = QWidget()
        card.setFixedWidth(380)
        card.setStyleSheet(
            "background:rgba(13,21,32,0.92);border:1px solid #172338;"
            "border-top:1px solid rgba(56,189,248,0.4);border-radius:16px;")
        cl = QVBoxLayout(card)
        cl.setContentsMargins(28, 26, 28, 24)
        cl.setSpacing(10)

        logo = QLabel("TEAM ANALYTICS")
        logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo.setStyleSheet(
            f"background:transparent;color:#38bdf8;font-size:16px;font-weight:800;"
            f"letter-spacing:3px;border:none;font-family:{MONO};")
        cl.addWidget(logo)

        sub = QLabel("POS Support  vs  POS EFT")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:11px;border:none;")
        cl.addWidget(sub)
        cl.addSpacing(6)

        hint = QLabel("Required scopes: Request(Read) · Person(Read) · Team(Read)")
        hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hint.setWordWrap(True)
        hint.setStyleSheet(
            "background:transparent;color:#6b83a0;font-size:10px;border:none;")
        cl.addWidget(hint)
        cl.addSpacing(8)

        self._tok_in = QLineEdit()
        self._tok_in.setPlaceholderText("Personal Access Token")
        self._tok_in.setEchoMode(QLineEdit.EchoMode.Password)
        self._tok_in.setFixedHeight(40)
        self._tok_in.setStyleSheet(self._input_qss())
        cl.addWidget(self._tok_in)

        self._acc_in = QLineEdit()
        self._acc_in.setPlaceholderText("Account ID  (pnp-it-services)")
        self._acc_in.setText("pnp-it-services")
        self._acc_in.setFixedHeight(40)
        self._acc_in.setStyleSheet(self._input_qss())
        cl.addWidget(self._acc_in)

        self._login_err = QLabel("")
        self._login_err.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._login_err.setWordWrap(True)
        self._login_err.setStyleSheet(
            "background:transparent;color:#f87171;font-size:10px;border:none;")
        cl.addWidget(self._login_err)

        conn = QPushButton("Connect to 4me")
        conn.setFixedHeight(42)
        conn.setCursor(Qt.CursorShape.PointingHandCursor)
        conn.setStyleSheet(
            "QPushButton{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #38bdf8,stop:1 #a78bfa);color:#060a10;border:none;"
            "border-radius:12px;font-size:13px;font-weight:700;letter-spacing:1px;}"
            "QPushButton:hover{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #5cc9ff,stop:1 #b99cff);}"
            "QPushButton:pressed{padding-top:1px;}")
        conn.clicked.connect(self._on_connect)
        cl.addWidget(conn)

        wl.addWidget(card, alignment=Qt.AlignmentFlag.AlignCenter)
        wl.addStretch()
        self._root.addWidget(wrap)

    def _input_qss(self):
        return ("QLineEdit{background:#070c16;border:1px solid #1f2d3d;"
                "border-radius:8px;color:#d4dfe9;font-size:12px;padding:0 14px;}"
                "QLineEdit:focus{border-color:#38bdf8;}"
                "QLineEdit::placeholder{color:#6b83a0;}")

    def _build_dashboard(self):
        from PyQt6.QtWidgets import QTabWidget
        tabs = QTabWidget()
        tabs.setStyleSheet(
            "QTabWidget::pane{border:none;}"
            "QTabBar::tab{background:transparent;color:#6d84a0;padding:8px 16px;"
            "font-size:11px;font-weight:700;border:none;border-bottom:2px solid transparent;}"
            "QTabBar::tab:selected{color:#38bdf8;border-bottom:2px solid #38bdf8;}"
            "QTabBar::tab:hover{color:#a7b6c8;}")
        overview_page = QWidget()
        self._overview_layout = QVBoxLayout(overview_page)
        self._overview_layout.setContentsMargins(4, 12, 4, 4)
        self._overview_layout.setSpacing(16)
        tabs.addTab(overview_page, "Overview")
        self._root.addWidget(tabs)
        self._build_store_till_tab(tabs)

        # ── HERO BAND ──────────────────────────────────────────────
        hero = QWidget()
        hero.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        hero.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:1,"
            "stop:0 rgba(56,189,248,0.10),stop:0.5 rgba(13,21,32,0.55),"
            "stop:1 rgba(167,139,250,0.10));"
            "border:1px solid #1c2b42;border-top:1px solid rgba(255,255,255,0.10);"
            "border-radius:16px;")
        hv = QVBoxLayout(hero)
        hv.setContentsMargins(22, 16, 22, 16)
        hv.setSpacing(4)

        matchup = QHBoxLayout()
        matchup.setSpacing(0)
        a_name = QLabel("POS SUPPORT")
        a_name.setStyleSheet(
            f"background:transparent;color:{TEAM_COLORS['POS Support']};"
            f"font-size:20px;font-weight:900;letter-spacing:1px;border:none;"
            f"font-family:{MONO};")
        vs = QLabel("  vs  ")
        vs.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:14px;"
            "font-weight:700;border:none;")
        b_name = QLabel("POS EFT")
        b_name.setStyleSheet(
            f"background:transparent;color:{TEAM_COLORS['POS EFT']};"
            f"font-size:20px;font-weight:900;letter-spacing:1px;border:none;"
            f"font-family:{MONO};")
        matchup.addWidget(a_name)
        matchup.addWidget(vs)
        matchup.addWidget(b_name)
        matchup.addStretch()

        self._live_pill = QLabel("● LIVE")
        self._live_pill.setStyleSheet(
            "background:rgba(0,232,122,0.12);color:#00e87a;font-size:10px;"
            "font-weight:800;letter-spacing:1px;border:1px solid rgba(0,232,122,0.3);"
            "border-radius:8px;padding:3px 10px;")
        matchup.addWidget(self._live_pill)
        hv.addLayout(matchup)

        tagline = QLabel("Head-to-head performance · auto-refreshes every 3 minutes")
        tagline.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;border:none;")
        hv.addWidget(tagline)

        # ── Open-call counters (live count of unresolved requests per team) ──
        # Click a badge to expand its per-status breakdown.
        open_row = QHBoxLayout()
        open_row.setSpacing(10)
        open_row.setContentsMargins(0, 8, 0, 0)

        self._open_breakdown_data = {}   # team -> {status: count}
        self._open_breakdown_panels = {}
        self._open_expanded = {"POS Support": False, "POS EFT": False}

        def _open_badge(team_key):
            color = TEAM_COLORS[team_key]
            box = _ClickableBox()
            box.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
            box.setCursor(Qt.CursorShape.PointingHandCursor)
            box.setStyleSheet(
                f"background:rgba(255,255,255,0.03);"
                f"border:1px solid {color}44;border-radius:10px;")
            bl = QHBoxLayout(box)
            bl.setContentsMargins(12, 6, 12, 6); bl.setSpacing(8)
            dot = QLabel("●")
            dot.setStyleSheet(f"background:transparent;color:{color};"
                              f"font-size:9px;border:none;")
            name = QLabel(team_key)
            name.setStyleSheet(
                f"background:transparent;color:{color};font-size:11px;"
                f"font-weight:800;letter-spacing:0.5px;border:none;")
            count = QLabel("—")
            count.setStyleSheet(
                f"background:transparent;color:#e8eef5;font-size:18px;"
                f"font-weight:900;border:none;font-family:{MONO};")
            lbl = QLabel("open")
            lbl.setStyleSheet(
                "background:transparent;color:#6d84a0;font-size:10px;"
                "font-weight:600;border:none;")
            chevron = QLabel("▾")
            chevron.setStyleSheet(
                "background:transparent;color:#4a5b70;font-size:10px;border:none;")
            bl.addWidget(dot); bl.addWidget(name)
            bl.addStretch()
            bl.addWidget(count); bl.addWidget(lbl); bl.addWidget(chevron)
            box.clicked.connect(lambda t=team_key: self._toggle_open_breakdown(t))
            return box, count, chevron

        sup_box, self._open_sup, self._open_sup_chev = _open_badge("POS Support")
        eft_box, self._open_eft, self._open_eft_chev = _open_badge("POS EFT")
        open_row.addWidget(sup_box, 1)
        open_row.addWidget(eft_box, 1)
        hv.addLayout(open_row)

        # Breakdown panels — hidden until their badge is clicked.
        panel_row = QHBoxLayout()
        panel_row.setSpacing(10)
        for team_key in ("POS Support", "POS EFT"):
            panel = QWidget()
            panel.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
            panel.setStyleSheet(
                "background:rgba(255,255,255,0.02);border:1px solid #1c2b42;"
                "border-radius:10px;")
            panel.setVisible(False)
            pl = QVBoxLayout(panel)
            pl.setContentsMargins(12, 8, 12, 8); pl.setSpacing(4)
            self._open_breakdown_panels[team_key] = (panel, pl)
            panel_row.addWidget(panel, 1)
        hv.addLayout(panel_row)

        self._overview_layout.addWidget(hero)

        # ── control row: period toggle + status + refresh ──
        header = QHBoxLayout()
        header.setSpacing(8)

        self._period_btns = {}
        for key, label in [("today", "Today"), ("week", "This Week"), ("month", "This Month")]:
            b = QPushButton(label)
            b.setCursor(Qt.CursorShape.PointingHandCursor)
            b.setFixedHeight(30)
            b.clicked.connect(lambda _, k=key: self._set_period(k))
            self._period_btns[key] = b
            header.addWidget(b)
        header.addStretch()

        self._status = QLabel("")
        self._status.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;border:none;")
        header.addWidget(self._status)

        export = QPushButton("⤓ CSV")
        export.setFixedHeight(30)
        export.setCursor(Qt.CursorShape.PointingHandCursor)
        export.setToolTip("Export the current period's data to a CSV file")
        export.setStyleSheet(
            "QPushButton{background:rgba(255,255,255,0.05);color:#8ba0b6;"
            "border:1px solid #1f2d3d;border-radius:8px;font-size:11px;"
            "font-weight:600;padding:0 12px;}"
            "QPushButton:hover{background:rgba(167,139,250,0.15);color:#a78bfa;"
            "border-color:#a78bfa;}")
        export.clicked.connect(self._export_csv)
        header.addSpacing(6)
        header.addWidget(export)

        refresh = QPushButton("⟳")
        refresh.setFixedSize(30, 30)
        refresh.setCursor(Qt.CursorShape.PointingHandCursor)
        refresh.setToolTip("Refresh now  (R)")
        refresh.setStyleSheet(
            "QPushButton{background:rgba(255,255,255,0.05);color:#8ba0b6;"
            "border:1px solid #1f2d3d;border-radius:8px;font-size:14px;}"
            "QPushButton:hover{background:rgba(56,189,248,0.15);color:#38bdf8;"
            "border-color:#38bdf8;}")
        refresh.clicked.connect(self._fetch)
        header.addSpacing(6)
        header.addWidget(refresh)

        self._overview_layout.addLayout(header)
        self._style_period_btns()

        # ── at-a-glance: Today / Week / Month totals for both teams ──
        self._glance = PeriodGlance()
        self._overview_layout.addWidget(self._glance)

        # ── head-to-head scoreboard (the verdict strip) ──
        self._scoreboard = Scoreboard()
        self._overview_layout.addWidget(self._scoreboard)

        # ── two team cards side by side ──
        cards_row = QHBoxLayout()
        cards_row.setSpacing(12)
        self._team_cards = {}
        for team in TEAM_IDS:
            c = TeamCard(team)
            self._team_cards[team] = c
            cards_row.addWidget(c, 1)
        self._overview_layout.addLayout(cards_row, 1)

        # ── trend line ──
        trend_wrap = QWidget()
        trend_wrap.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        trend_wrap.setStyleSheet(
            "background:rgba(13,21,32,0.6);border:1px solid #172338;"
            "border-radius:12px;")
        tw = QVBoxLayout(trend_wrap)
        tw.setContentsMargins(16, 12, 16, 12)
        tw.setSpacing(6)
        tl = QLabel("DAILY CLOSES — LAST 14 DAYS")
        tl.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;"
            "font-weight:700;letter-spacing:1.5px;border:none;")
        self._trend_title = tl
        tw.addWidget(tl)
        self._trend = TrendLine()
        self._trend.setMinimumHeight(220)
        tw.addWidget(self._trend)
        self._overview_layout.addWidget(trend_wrap)

        # ── cross-team handoffs — compact one-liner ──
        self._handoff = QLabel("")
        self._handoff.setWordWrap(True)
        self._handoff.setStyleSheet(
            "background:rgba(13,21,32,0.4);border:1px solid #172338;"
            "border-radius:12px;color:#8ba0b6;font-size:11px;padding:10px 14px;")
        self._handoff.setToolTip(
            "Reassignments — closed tickets that changed teams before completing "
            "(4me assignmentCount > 1). The API exposes how many times a ticket "
            "was reassigned, not the direction, so this shows which team received "
            "reassigned work, not who sent it.")
        self._overview_layout.addWidget(self._handoff)

        # sign-out (small)
        foot = QHBoxLayout()
        foot.addStretch()
        out = QPushButton("disconnect")
        out.setCursor(Qt.CursorShape.PointingHandCursor)
        out.setStyleSheet(
            "QPushButton{background:transparent;border:none;color:#6b83a0;"
            "font-size:10px;} QPushButton:hover{color:#f87171;}")
        out.clicked.connect(self._disconnect)
        foot.addWidget(out)
        self._overview_layout.addLayout(foot)

    # ── period + styling ─────────────────────────────────────────
    def _set_period(self, key):
        self._period = key
        self._last_totals = {}   # delta is per-period, reset on switch
        self._style_period_btns()
        # instant render from cache if we have it, then refresh live
        cached = self._period_cache.get(key)
        if cached:
            self._on_data(cached, key, primary=True)
        # switch the trend window/title right away (uses month cache)
        if hasattr(self, "_trend"):
            self._refresh_trend()
        self._fetch()

    def _build_store_till_tab(self, tabs):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(4, 12, 4, 4)
        layout.setSpacing(14)

        title = QLabel("STORE & TILL BREAKDOWN")
        title.setStyleSheet(
            "background:transparent;color:#8ba0b6;font-size:11px;"
            "font-weight:800;letter-spacing:1.5px;border:none;")
        layout.addWidget(title)
        sub = QLabel(
            "Classifies every call by store format (Hyper / Express / Local "
            "/ Liquor / Clothing, read from the requester's name) and till "
            "parity (from the Till Number field). Pick a team and period, "
            "then load and export.")
        sub.setWordWrap(True)
        sub.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;border:none;")
        layout.addWidget(sub)

        controls = QHBoxLayout()
        controls.setSpacing(8)

        self._bd_team = "POS Support"
        self._bd_team_btns = {}
        for team in TEAM_IDS:
            b = QPushButton(team)
            b.setCursor(Qt.CursorShape.PointingHandCursor)
            b.setFixedHeight(30)
            b.clicked.connect(lambda _, t=team: self._set_breakdown_team(t))
            self._bd_team_btns[team] = b
            controls.addWidget(b)

        controls.addSpacing(12)

        self._bd_period = "month"
        self._bd_period_btns = {}
        for key, label in [("today", "Today"), ("week", "This Week"), ("month", "This Month")]:
            b = QPushButton(label)
            b.setCursor(Qt.CursorShape.PointingHandCursor)
            b.setFixedHeight(30)
            b.clicked.connect(lambda _, k=key: self._set_breakdown_period(k))
            self._bd_period_btns[key] = b
            controls.addWidget(b)

        controls.addStretch()

        load_btn = QPushButton("⟳  Load")
        load_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        load_btn.setFixedHeight(30)
        load_btn.clicked.connect(self._fetch_calls_breakdown)
        load_btn.setStyleSheet(
            "QPushButton{background:rgba(56,189,248,0.12);color:#38bdf8;"
            "border:1px solid rgba(56,189,248,0.35);border-radius:8px;"
            "padding:0 14px;font-size:11px;font-weight:700;}"
            "QPushButton:hover{background:rgba(56,189,248,0.2);}"
            "QPushButton:disabled{color:#3a4a5c;border-color:#1c2b42;}")
        controls.addWidget(load_btn)
        self._bd_load_btn = load_btn

        layout.addLayout(controls)

        self._bd_status = QLabel(
            "Click Load to fetch call data for the selected team and period.")
        self._bd_status.setWordWrap(True)
        self._bd_status.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;border:none;")
        layout.addWidget(self._bd_status)

        # ── filters: narrow both the summary AND the export ──
        self._bd_all_formats = ["Hyper", "Express", "Local", "Liquor", "Clothing", "Other", "Unknown"]
        self._bd_all_parities = ["Odd", "Even", "Unknown"]
        self._bd_format_filter = set(self._bd_all_formats)   # all active by default
        self._bd_parity_filter = set(self._bd_all_parities)
        self._bd_format_chips = {}
        self._bd_parity_chips = {}

        filter_wrap = QWidget()
        fw = QVBoxLayout(filter_wrap)
        fw.setContentsMargins(0, 4, 0, 4); fw.setSpacing(6)

        fmt_row = QHBoxLayout(); fmt_row.setSpacing(6)
        fmt_lbl = QLabel("Format:")
        fmt_lbl.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;"
            "font-weight:700;border:none;")
        fmt_row.addWidget(fmt_lbl)
        for fmt in self._bd_all_formats:
            chip = QPushButton(fmt)
            chip.setCheckable(True)
            chip.setChecked(True)
            chip.setCursor(Qt.CursorShape.PointingHandCursor)
            chip.setFixedHeight(24)
            chip.clicked.connect(lambda _, f=fmt: self._toggle_bd_filter("format", f))
            self._bd_format_chips[fmt] = chip
            fmt_row.addWidget(chip)
        fmt_row.addStretch()
        fw.addLayout(fmt_row)

        par_row = QHBoxLayout(); par_row.setSpacing(6)
        par_lbl = QLabel("Till parity:")
        par_lbl.setStyleSheet(
            "background:transparent;color:#6d84a0;font-size:10px;"
            "font-weight:700;border:none;")
        par_row.addWidget(par_lbl)
        for parity in self._bd_all_parities:
            chip = QPushButton(parity)
            chip.setCheckable(True)
            chip.setChecked(True)
            chip.setCursor(Qt.CursorShape.PointingHandCursor)
            chip.setFixedHeight(24)
            chip.clicked.connect(lambda _, p=parity: self._toggle_bd_filter("parity", p))
            self._bd_parity_chips[parity] = chip
            par_row.addWidget(chip)
        par_row.addStretch()
        clear_btn = QPushButton("Clear filters")
        clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        clear_btn.setFixedHeight(24)
        clear_btn.clicked.connect(self._clear_bd_filters)
        clear_btn.setStyleSheet(
            "QPushButton{background:transparent;color:#6d84a0;border:1px solid #1f2d3d;"
            "border-radius:8px;padding:0 10px;font-size:10px;}"
            "QPushButton:hover{color:#d4dfe9;border-color:#6b83a0;}")
        par_row.addWidget(clear_btn)
        fw.addLayout(par_row)

        layout.addWidget(filter_wrap)
        self._style_bd_chips()

        # ── summary: format counts + parity counts side by side ──
        summary_row = QHBoxLayout()
        summary_row.setSpacing(12)

        def _summary_card(title_text):
            card = QWidget()
            card.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
            card.setStyleSheet(
                "background:rgba(255,255,255,0.02);border:1px solid #1c2b42;"
                "border-radius:12px;")
            cl = QVBoxLayout(card)
            cl.setContentsMargins(14, 12, 14, 12); cl.setSpacing(6)
            t = QLabel(title_text)
            t.setStyleSheet(
                "background:transparent;color:#6d84a0;font-size:10px;"
                "font-weight:800;letter-spacing:1px;border:none;")
            cl.addWidget(t)
            inner = QVBoxLayout(); inner.setSpacing(4)
            cl.addLayout(inner)
            return card, inner

        fmt_card, self._bd_fmt_layout = _summary_card("BY FORMAT")
        par_card, self._bd_par_layout = _summary_card("BY TILL PARITY")
        summary_row.addWidget(fmt_card, 1)
        summary_row.addWidget(par_card, 1)
        layout.addLayout(summary_row)

        export_btn = QPushButton("⬇  Export to Excel")
        export_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        export_btn.setFixedHeight(34)
        export_btn.setEnabled(False)
        export_btn.clicked.connect(self._export_calls_breakdown_xlsx)
        export_btn.setStyleSheet(
            "QPushButton{background:rgba(0,232,122,0.12);color:#00e87a;"
            "border:1px solid rgba(0,232,122,0.35);border-radius:8px;"
            "font-size:12px;font-weight:700;}"
            "QPushButton:disabled{color:#3a4a5c;border-color:#1c2b42;"
            "background:transparent;}"
            "QPushButton:hover:!disabled{background:rgba(0,232,122,0.2);}")
        layout.addWidget(export_btn)
        self._bd_export_btn = export_btn

        layout.addStretch()
        tabs.addTab(page, "Store & Till")

        self._calls_breakdown_data = {}
        self._style_breakdown_btns()

    def _style_bd_chips(self):
        def _style_chip(chip, active):
            if active:
                chip.setStyleSheet(
                    "QPushButton{background:rgba(56,189,248,0.18);color:#38bdf8;"
                    "border:1px solid #38bdf8;border-radius:8px;padding:0 10px;"
                    "font-size:10px;font-weight:700;}")
            else:
                chip.setStyleSheet(
                    "QPushButton{background:transparent;color:#4a5b70;"
                    "border:1px solid #1c2b42;border-radius:8px;padding:0 10px;"
                    "font-size:10px;}"
                    "QPushButton:hover{color:#6d84a0;}")
        for fmt, chip in self._bd_format_chips.items():
            _style_chip(chip, fmt in self._bd_format_filter)
        for parity, chip in self._bd_parity_chips.items():
            _style_chip(chip, parity in self._bd_parity_filter)

    def _toggle_bd_filter(self, kind, value):
        target = self._bd_format_filter if kind == "format" else self._bd_parity_filter
        if value in target:
            target.discard(value)
        else:
            target.add(value)
        self._style_bd_chips()
        self._render_breakdown_summary()

    def _clear_bd_filters(self):
        self._bd_format_filter = set(self._bd_all_formats)
        self._bd_parity_filter = set(self._bd_all_parities)
        for chip in list(self._bd_format_chips.values()) + list(self._bd_parity_chips.values()):
            chip.setChecked(True)
        self._style_bd_chips()
        self._render_breakdown_summary()

    def _filtered_calls(self):
        """The selected team's calls, narrowed by the active format/parity
        filter chips. Both the on-screen summary and the export use this same
        list, so what you see is exactly what you'd get in the spreadsheet."""
        data = getattr(self, "_calls_breakdown_data", {}).get(self._bd_team)
        if not data:
            return []
        return [c for c in data["calls"]
                if c["format"] in self._bd_format_filter
                and c["parity"] in self._bd_parity_filter]

    def _style_breakdown_btns(self):
        for team, b in getattr(self, "_bd_team_btns", {}).items():
            color = TEAM_COLORS[team]
            if team == self._bd_team:
                b.setStyleSheet(
                    f"QPushButton{{background:{color};color:#060a10;"
                    f"border:none;border-radius:8px;padding:0 14px;"
                    f"font-size:11px;font-weight:700;}}")
            else:
                b.setStyleSheet(
                    "QPushButton{background:transparent;color:#6d84a0;"
                    "border:1px solid #1f2d3d;border-radius:8px;"
                    "padding:0 14px;font-size:11px;}"
                    "QPushButton:hover{color:#d4dfe9;border-color:#6b83a0;}")
        for key, b in getattr(self, "_bd_period_btns", {}).items():
            if key == self._bd_period:
                b.setStyleSheet(
                    "QPushButton{background:qlineargradient(x1:0,y1:0,x2:1,y2:1,"
                    "stop:0 #38bdf8,stop:1 #a78bfa);color:#060a10;border:none;"
                    "border-radius:8px;padding:0 14px;font-size:10px;font-weight:700;}")
            else:
                b.setStyleSheet(
                    "QPushButton{background:transparent;color:#6d84a0;"
                    "border:1px solid #1f2d3d;border-radius:8px;"
                    "padding:0 14px;font-size:10px;}"
                    "QPushButton:hover{color:#d4dfe9;border-color:#6b83a0;}")

    def _set_breakdown_team(self, team):
        self._bd_team = team
        self._style_breakdown_btns()
        self._render_breakdown_summary()

    def _set_breakdown_period(self, key):
        self._bd_period = key
        self._style_breakdown_btns()
        self._bd_export_btn.setEnabled(False)
        self._bd_status.setText("Period changed — click Load to refresh.")

    def _fetch_calls_breakdown(self):
        if not self._token or not self._account:
            return
        from PyQt6.QtCore import QThread, pyqtSignal

        class _BreakdownWorker(QThread):
            done = pyqtSignal(dict)
            error = pyqtSignal(str)
            def __init__(self, token, account, start_date):
                super().__init__()
                self.token = token; self.account = account
                self.start_date = start_date
            def run(self):
                try:
                    from src.widgets.xurrent_client import fetch_calls_breakdown
                    self.done.emit(
                        fetch_calls_breakdown(self.token, self.account, self.start_date))
                except Exception as e:
                    self.error.emit(str(e))

        self._bd_load_btn.setEnabled(False)
        self._bd_status.setText("Loading — this can take a moment for a full month…")
        start_date = self._start_date(self._bd_period)
        self._bd_worker = _BreakdownWorker(self._token, self._account, start_date)
        self._bd_worker.done.connect(self._on_calls_breakdown)
        self._bd_worker.error.connect(self._on_calls_breakdown_error)
        self._bd_worker.start()

    def _on_calls_breakdown(self, data):
        self._calls_breakdown_data = data
        self._bd_load_btn.setEnabled(True)
        total = sum(v.get("total", 0) for v in data.values())
        self._bd_status.setText(f"Loaded {total} calls across both teams.")
        self._bd_export_btn.setEnabled(True)
        self._render_breakdown_summary()

    def _on_calls_breakdown_error(self, msg):
        self._bd_load_btn.setEnabled(True)
        short = (msg[:100] + "…") if len(msg) > 100 else msg
        self._bd_status.setText(f"Couldn't load: {short}")
        self._bd_export_btn.setEnabled(False)

    def _render_breakdown_summary(self):
        for lay in (getattr(self, "_bd_fmt_layout", None), getattr(self, "_bd_par_layout", None)):
            if lay is None:
                continue
            while lay.count():
                item = lay.takeAt(0)
                if item and item.widget():
                    item.widget().deleteLater()

        if self._bd_team not in getattr(self, "_calls_breakdown_data", {}):
            return
        calls = self._filtered_calls()
        color = TEAM_COLORS[self._bd_team]

        format_counts = {}
        parity_counts = {}
        for c in calls:
            format_counts[c["format"]] = format_counts.get(c["format"], 0) + 1
            parity_counts[c["parity"]] = parity_counts.get(c["parity"], 0) + 1

        for fmt in ("Hyper", "Express", "Local", "Liquor", "Clothing", "Other", "Unknown"):
            n = format_counts.get(fmt, 0)
            if n:
                self._add_breakdown_row(self._bd_fmt_layout, fmt, n, color)
        for parity in ("Odd", "Even", "Unknown"):
            n = parity_counts.get(parity, 0)
            if n:
                self._add_breakdown_row(self._bd_par_layout, parity, n, color)

        if not calls:
            empty = QLabel("No calls match the current filters.")
            empty.setStyleSheet(
                "background:transparent;color:#6d84a0;font-size:10px;border:none;")
            self._bd_fmt_layout.addWidget(empty)

    def _export_calls_breakdown_xlsx(self):
        data = getattr(self, "_calls_breakdown_data", {}).get(self._bd_team)
        if not data:
            self._bd_status.setText("Nothing to export yet — click Load first.")
            return
        calls = self._filtered_calls()
        if not calls:
            self._bd_status.setText(
                "No calls match the current filters — nothing to export.")
            return
        from PyQt6.QtWidgets import QFileDialog
        import datetime
        default_name = (f"{self._bd_team.replace(' ', '_')}_calls_"
                        f"{datetime.date.today().isoformat()}.xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self, "Export calls breakdown", default_name, "Excel files (*.xlsx)")
        if not path:
            return
        try:
            self._write_breakdown_xlsx(path, calls)
            self._bd_status.setText(
                f"Exported {len(calls)} rows to {path}")
        except Exception as e:
            self._bd_status.setText(f"Export failed: {e}")

    def _write_breakdown_xlsx(self, path, calls):
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Calls"
        ws.append(["Request Number", "Store Code", "Till Number", "Odd/Even", "Format"])
        for c in ws[1]:
            c.font = Font(bold=True, name="Arial")
        for call in calls:
            ws.append([
                call["request_id"],
                call["store_code"],
                call["till_number"] if call["till_number"] is not None else "",
                call["parity"],
                call["format"],
            ])
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial")
        for i, w in enumerate([16, 12, 12, 10, 12], start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        wb.save(path)

    def _style_period_btns(self):
        if not hasattr(self, "_period_btns"):
            return
        for key, b in self._period_btns.items():
            if key == self._period:
                b.setStyleSheet(
                    "QPushButton{background:qlineargradient(x1:0,y1:0,x2:1,y2:1,"
                    "stop:0 #38bdf8,stop:1 #a78bfa);color:#060a10;border:none;"
                    "border-radius:8px;padding:0 14px;font-size:10px;font-weight:700;}")
            else:
                b.setStyleSheet(
                    "QPushButton{background:transparent;color:#6d84a0;"
                    "border:1px solid #1f2d3d;border-radius:8px;padding:0 14px;"
                    "font-size:10px;font-weight:500;}"
                    "QPushButton:hover{color:#d4dfe9;border-color:#6b83a0;}")

    _PERIOD_FN = {"today": today_str, "week": week_start, "month": month_start}

    def _start_date(self, period=None):
        return self._PERIOD_FN[period or self._period]()

    # ── fetch + render ───────────────────────────────────────────
    def _fetch(self):
        """Refresh the active period (full render) + the other two periods
        in the background so the at-a-glance strip stays populated."""
        if not self._token or not self._account:
            return
        if hasattr(self, "_status"):
            self._status.setText("refreshing…")
        self._spawn_worker(self._period, primary=True)
        for pk in ("today", "week", "month"):
            if pk != self._period:
                self._spawn_worker(pk, primary=False)
        self._fetch_open_counts()

    def _fetch_open_counts(self):
        """Fetch the live open-request count per team (lightweight, its own
        thread so it never blocks the main analytics render)."""
        from PyQt6.QtCore import QThread, pyqtSignal

        class _OpenWorker(QThread):
            done = pyqtSignal(dict)
            error = pyqtSignal(str)
            def __init__(self, token, account):
                super().__init__(); self.token = token; self.account = account
            def run(self):
                try:
                    from src.widgets.xurrent_client import fetch_open_counts
                    self.done.emit(fetch_open_counts(self.token, self.account))
                except Exception as e:
                    self.error.emit(str(e))

        if getattr(self, "_open_worker", None) and self._open_worker.isRunning():
            return
        self._open_worker = _OpenWorker(self._token, self._account)
        self._open_worker.done.connect(self._on_open_counts)
        self._open_worker.error.connect(self._on_open_error)
        self._open_worker.start()

    def _on_open_error(self, msg):
        """
        Surface the real error instead of leaving the badges blank forever.
        Previously this was silently swallowed, which made a broken open-count
        query indistinguishable from 'still loading' — showing the actual
        message (truncated) makes it diagnosable.
        """
        short = (msg[:40] + "…") if len(msg) > 40 else msg
        for lbl in (getattr(self, "_open_sup", None), getattr(self, "_open_eft", None)):
            if lbl:
                lbl.setText("!")
                lbl.setToolTip(f"Couldn't load open-call count:\n{msg}")
        if hasattr(self, "_status"):
            self._status.setText(f"open-count error: {short}")

    def _toggle_open_breakdown(self, team_key):
        panel, _ = self._open_breakdown_panels[team_key]
        expanded = not self._open_expanded[team_key]
        self._open_expanded[team_key] = expanded
        chev = self._open_sup_chev if team_key == "POS Support" else self._open_eft_chev
        chev.setText("▴" if expanded else "▾")
        panel.setVisible(expanded)
        if expanded:
            self._render_open_breakdown(team_key)

    def _add_breakdown_row(self, layout, label, count, value_color):
        row_w = QWidget()
        row = QHBoxLayout(row_w)
        row.setContentsMargins(0, 0, 0, 0); row.setSpacing(8)
        lbl = QLabel(label)
        lbl.setStyleSheet(
            "background:transparent;color:#a7b6c8;font-size:11px;border:none;")
        val = QLabel(str(count))
        val.setStyleSheet(
            f"background:transparent;color:{value_color};font-size:11px;"
            f"font-weight:800;border:none;font-family:{MONO};")
        row.addWidget(lbl); row.addStretch(); row.addWidget(val)
        layout.addWidget(row_w)

    def _render_open_breakdown(self, team_key):
        panel, layout = self._open_breakdown_panels[team_key]
        while layout.count():
            item = layout.takeAt(0)
            if item and item.widget():
                item.widget().deleteLater()

        data = self._open_breakdown_data.get(team_key)
        if not data:
            empty = QLabel("No breakdown yet — refreshing…")
            empty.setStyleSheet(
                "background:transparent;color:#6d84a0;font-size:10px;border:none;")
            layout.addWidget(empty)
            return

        breakdown = data.get("breakdown", {})
        color = TEAM_COLORS[team_key]
        any_rows = False
        for status in STATUS_ORDER:
            bucket = breakdown.get(status, {"member": 0, "unclaimed": 0})
            member_n = bucket.get("member", 0)
            unclaimed_n = bucket.get("unclaimed", 0)
            if member_n == 0 and unclaimed_n == 0:
                continue
            any_rows = True
            label = STATUS_LABELS[status]
            # Only split into two rows when this status actually has BOTH
            # member-assigned and unclaimed requests in the real data — a
            # status that's uniformly one or the other just gets one row.
            if member_n > 0:
                any_rows = True
                self._add_breakdown_row(layout, label, member_n, color)
            if unclaimed_n > 0:
                any_rows = True
                # Amber-tinted: unclaimed work needs someone to pick it up.
                self._add_breakdown_row(
                    layout, f"{label} (Unclaimed)", unclaimed_n, "#fbbf24")

        if not any_rows:
            empty = QLabel("Nothing open right now.")
            empty.setStyleSheet(
                "background:transparent;color:#6d84a0;font-size:10px;border:none;")
            layout.addWidget(empty)

    def _on_open_counts(self, counts):
        """counts = {team_name: {"total": int, "breakdown": {...}}} — exact,
        fully paginated (see fetch_open_counts)."""
        for team_key, lbl in (("POS Support", getattr(self, "_open_sup", None)),
                              ("POS EFT", getattr(self, "_open_eft", None))):
            if not lbl:
                continue
            data = counts.get(team_key)
            if data is None:
                continue
            self._open_breakdown_data[team_key] = data
            lbl.setText(str(data.get("total", "—")))
            lbl.setToolTip("")
            if self._open_expanded.get(team_key):
                self._render_open_breakdown(team_key)

    def _spawn_worker(self, period, primary):
        if not hasattr(self, "_workers"):
            self._workers = {}
        existing = self._workers.get(period)
        if existing and existing.isRunning():
            return
        w = AnalyticsWorker(self._token, self._account, self._start_date(period))
        w.done.connect(lambda data, p=period, pr=primary: self._on_data(data, p, pr))
        w.error.connect(self._on_error)
        self._workers[period] = w
        w.start()

    def _export_csv(self):
        """Write the current period's per-agent + team metrics to a CSV."""
        from PyQt6.QtWidgets import QFileDialog
        import csv, datetime
        data = self._period_cache.get(self._period)
        if not data:
            if hasattr(self, "_status"):
                self._status.setText("nothing to export yet")
            return
        default = f"team_analytics_{self._period}_{datetime.date.today().isoformat()}.csv"
        path, _ = QFileDialog.getSaveFileName(self, "Export CSV", default, "CSV (*.csv)")
        if not path:
            return
        metrics = data.get("metrics", {})
        team_agent = data.get("team_agent", {})
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                wr = csv.writer(f)
                wr.writerow([f"Team Analytics — {self._period} — {datetime.datetime.now():%Y-%m-%d %H:%M}"])
                wr.writerow([])
                wr.writerow(["TEAM", "Closed", "Active", "Per-Agent",
                             "Median Hours", "Consistency", "Balance",
                             "Reassigned In", "Reopened"])
                for team, m in metrics.items():
                    wr.writerow([team, m.get("total", 0), m.get("active", 0),
                                 round(m.get("per_agent", 0), 1),
                                 round(m["med_hours"], 1) if m.get("med_hours") is not None else "",
                                 round(m["steadiness"]) if m.get("steadiness") is not None else "",
                                 round(m["balance"]) if m.get("balance") is not None else "",
                                 m.get("reassigned_in", 0), m.get("reopened", 0)])
                wr.writerow([])
                wr.writerow(["TEAM", "Agent", "Closed"])
                for team, agents in team_agent.items():
                    for name, cnt in sorted(agents.items(), key=lambda kv: kv[1], reverse=True):
                        wr.writerow([team, name, cnt])
            if hasattr(self, "_status"):
                self._status.setText("exported ✓")
        except Exception as e:
            if hasattr(self, "_status"):
                self._status.setText(f"export failed: {e}")

    def keyPressEvent(self, e):
        """Keyboard shortcuts: R refresh, 1/2/3 periods, Ctrl+E export."""
        k = e.key()
        if k == Qt.Key.Key_R:
            self._fetch()
        elif k == Qt.Key.Key_1:
            self._set_period("today")
        elif k == Qt.Key.Key_2:
            self._set_period("week")
        elif k == Qt.Key.Key_3:
            self._set_period("month")
        elif k == Qt.Key.Key_E and (e.modifiers() & Qt.KeyboardModifier.ControlModifier):
            self._export_csv()
        else:
            super().keyPressEvent(e)

    def _on_data(self, data, period=None, primary=True):
        by_team    = data.get("by_team", {})
        team_agent = data.get("team_agent", {})
        by_day     = data.get("by_day", {})
        metrics    = data.get("metrics", {})
        raw        = data.get("raw", [])

        # always feed the at-a-glance strip for this period
        if period and hasattr(self, "_glance"):
            self._glance.set_period(period, by_team)
        self._period_cache[period or self._period] = data

        # The trend always reflects the widest (month) data — so when the
        # month background fetch lands, refresh the chart even if the active
        # tab is Today/Week.
        if period == "month" and hasattr(self, "_trend"):
            self._refresh_trend()

        # only the active period drives the full dashboard
        if not primary:
            return

        for team, card in self._team_cards.items():
            total  = by_team.get(team, 0)
            agents = team_agent.get(team, {})
            prev   = self._last_totals.get(team)
            delta_txt, up = "", None
            if prev is not None:
                diff = total - prev
                if diff != 0:
                    delta_txt = f"{abs(diff)} vs last refresh"
                    up = diff > 0
            m = metrics.get(team, {})
            card.set_value(total, agents, delta_txt, up, m)

        self._last_totals = dict(by_team)

        # head-to-head scoreboard verdict → drives the hero leader badge
        champ = None
        if hasattr(self, "_scoreboard"):
            champ = self._scoreboard.update_metrics(metrics, list(TEAM_IDS.keys()))
        if hasattr(self, "_live_pill"):
            if champ:
                col = TEAM_COLORS.get(champ, "#00e87a")
                short = "SUPPORT" if "Support" in champ else "EFT"
                self._live_pill.setText(f"◆ {short} LEADING")
                self._live_pill.setStyleSheet(
                    f"background:{col}22;color:{col};font-size:10px;"
                    f"font-weight:800;letter-spacing:1px;border:1px solid {col}55;"
                    f"border-radius:8px;padding:3px 10px;")
            else:
                self._live_pill.setText("◆ TIED")
                self._live_pill.setStyleSheet(
                    "background:rgba(139,160,182,0.12);color:#8ba0b6;font-size:10px;"
                    "font-weight:800;letter-spacing:1px;border:1px solid rgba(139,160,182,0.3);"
                    "border-radius:8px;padding:3px 10px;")

        # cross-team handoffs — compact one-liner (from assignmentCount)
        if hasattr(self, "_handoff"):
            teams = list(TEAM_IDS.keys())
            a, b = teams[0], teams[1]
            a_in = metrics.get(a, {}).get("reassigned_in", 0)
            b_in = metrics.get(b, {}).get("reassigned_in", 0)
            total = (a_in + b_in) or 1
            if a_in == b_in:
                self._handoff.setText(
                    f"↔  Reassignments: {a} and {b} each absorbed {a_in} "
                    f"tickets handed over between teams.")
            else:
                taker = a if a_in > b_in else b
                pct = round(100 * max(a_in, b_in) / total)
                col = TEAM_COLORS.get(taker, "#38bdf8")
                self._handoff.setText(
                    f"↔  Reassignments: {a} absorbed {a_in} · {b} absorbed {b_in}  "
                    f"—  {taker} takes {pct}% of work handed between teams.")
                self._handoff.setStyleSheet(
                    f"background:rgba(13,21,32,0.4);border:1px solid #172338;"
                    f"border-left:3px solid {col};border-radius:12px;"
                    f"color:#9fb0c4;font-size:11px;padding:10px 14px;")

        # trend — always rebuilt from widest data, independent of active tab
        self._refresh_trend()

        from datetime import datetime
        if hasattr(self, "_status"):
            total = data.get("total", 0)
            if total == 0 and primary:
                # Zero records back: help diagnose WHY (creds/scope/date/network)
                # rather than just sitting on "Waiting for data".
                start = self._start_date(period)
                self._status.setText(f"0 records for {period} (since {start})")
                self._status.setToolTip(
                    "The query succeeded but returned 0 completed tickets.\n"
                    "Common causes on a different machine:\n"
                    "• The 4me token in ~/.4me_pro_creds.json differs from your\n"
                    "  main laptop (wrong account, or missing Request read scope).\n"
                    "• System date/timezone is off, so the date range is empty.\n"
                    "• A network/proxy is blocking the GraphQL response.\n"
                    "Try: disconnect and reconnect with the same token you use\n"
                    "on your main laptop.")
                self._status.setStyleSheet(
                    "background:transparent;color:#fbbf24;font-size:10px;border:none;")
            else:
                self._status.setText(f"updated {datetime.now().strftime('%H:%M')}")
                self._status.setToolTip("")
                self._status.setStyleSheet(
                    "background:transparent;color:#6b83a0;font-size:10px;border:none;")

    def _refresh_trend(self):
        """
        Trend window follows the selected period:
          • Today / This Week → last 14 days (recent context)
          • This Month        → every day from the 1st to today
        Always built from the widest data we have (month cache) so it's
        populated regardless of which tab triggered the fetch. Days with no
        closes render as 0 (continuous axis, no gaps).
        """
        from collections import defaultdict
        from datetime import date, timedelta

        src = self._period_cache.get("month") or self._period_cache.get(self._period)
        raw = src.get("raw", []) if src else []

        per = defaultdict(lambda: defaultdict(int))
        for t in raw:
            day = (t.get("updatedAt") or "")[:10]
            team = (t.get("team") or {}).get("name", "Unknown")
            if day and team in TEAM_IDS:
                per[team][day] += 1

        today = date.today()
        if self._period == "month":
            # every day of the current month, 1st → today
            start = today.replace(day=1)
            span = (today - start).days
            window = [(start + timedelta(days=i)).isoformat() for i in range(span + 1)]
            title = f"DAILY CLOSES — {today.strftime('%B').upper()} (EVERY DAY)"
        else:
            window = [(today - timedelta(days=i)).isoformat() for i in range(13, -1, -1)]
            title = "DAILY CLOSES — LAST 14 DAYS"

        if hasattr(self, "_trend_title"):
            self._trend_title.setText(title)

        series = {}
        for team in TEAM_IDS:
            series[team] = [(d, per[team].get(d, 0)) for d in window]
        self._trend.set_data(series, self._palette.get("accent", "#00e87a"))

    def _build_trend_series(self, raw):
        # kept for compatibility; now delegates to the always-on trend
        self._refresh_trend()
        return {}

    def _on_error(self, msg):
        if hasattr(self, "_status"):
            # Show a short error in the status line, full text on hover.
            short = msg.strip().split("\n")[0][:60]
            self._status.setText(f"⚠ {short}")
            self._status.setToolTip(msg)
            self._status.setStyleSheet(
                "background:transparent;color:#f87171;font-size:10px;border:none;")
        print(f"[TeamAnalytics] query error: {msg}")   # also to crash.txt/console
        # auth failures → bounce back to login
        low = msg.lower()
        if "auth" in low or "token" in low or "401" in low or "unauthor" in low:
            self._build_ui()

    # ── connect / disconnect ─────────────────────────────────────
    def _on_connect(self):
        tok = self._tok_in.text().strip()
        acc = self._acc_in.text().strip()
        if not tok:
            self._login_err.setText("Token required.")
            return
        if not acc:
            self._login_err.setText("Account ID required.")
            return
        vault_save(tok, acc)
        self._build_ui()

    def _disconnect(self):
        vault_clear()
        self._token = self._account = ""
        self._last_totals = {}
        self._build_ui()

    def set_palette(self, p):
        self._palette = p or {}


class TeamAnalyticsPanel(QWidget):
    """
    Invisible stack placeholder. Its refresh() (called by app._nav_to when
    the Teams pill is clicked) opens the independent popout window, then
    snaps the main stack back to Commands so nothing blank shows behind it.
    Mirrors the SettingsPanel pattern.
    """
    def __init__(self, app):
        super().__init__()
        self.app  = app
        self._win = None
        self.setStyleSheet("background:transparent;")

    def refresh(self):
        if not self._win:
            self._win = TeamAnalyticsWindow(self.app)
            if getattr(self.app, "_palette", None):
                self._win.set_palette(self.app._palette)
        self._win.open_window()
        # If connected, kick a fetch and (re)start the 3-min timer, since
        # closing the window stops it.
        if self._win._token and self._win._account:
            self._win._timer.start()
            self._win._fetch()
        # Return the main window to Commands so the invisible placeholder
        # page never shows.
        from PyQt6.QtCore import QTimer as _QT
        _QT.singleShot(50, lambda: self.app._stack.setCurrentIndex(
            list(self.app._panels.keys()).index('commands')
            if hasattr(self.app, '_panels') else 0))

    def set_palette(self, p):
        if self._win:
            self._win.set_palette(p)
