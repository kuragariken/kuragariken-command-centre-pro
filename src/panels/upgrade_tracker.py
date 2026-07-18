"""
Upgrade Command Centre — a Command Centre Pro module for tracking POS till
upgrade rollouts across stores.

Replaces the manual multi-tab Excel workbook with a single live dataset:
paste a raw SSH cache-health dump to add/refresh stores, click through tills
to mark them upgraded/failed, and get an instant dashboard + one-click
summary you can paste straight into Teams/email for the team.

Integration into Command Centre Pro
------------------------------------
    from upgrade_tracker import UpgradeTrackerWindow

    class CommandCentrePro(QMainWindow):
        def _open_upgrade_tracker(self):
            if not hasattr(self, "_upgrade_win") or self._upgrade_win is None:
                self._upgrade_win = UpgradeTrackerWindow()
            self._upgrade_win.show_fullscreen()

Wire that up to a HoverCard button the same way Vault/Notepad/Settings are
launched today. The window is self-contained (own data file, own styling)
and doesn't require anything else from the host app.

Standalone
----------
    python upgrade_tracker.py

Data lives at %APPDATA%\\Command Centre Pro\\UpgradeTracker.json on Windows
(falls back to ~/.command_centre_pro/UpgradeTracker.json elsewhere). On first
run ever (no data file present yet), if seed_data.json sits next to this
script, it's offered as a one-click import of your existing tracking data.
"""

import sys, os, re, json, csv, io, datetime, math
from datetime import datetime as dt

from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from PyQt6.QtGui import *

APP_VERSION = "1.0"

# ── Design tokens — shared aurora/glass language with CCP + 4me Pro ─────────
BG0   = "#05060c"
BG1   = "#0a0c16"
BG2   = "#111326"
BG3   = "#181b2e"
BG4   = "#1e2235"
SEP   = "#252840"

CY    = "#00d4ff"
CY2   = "#0088bb"
VIO   = "#7c3aed"
VIO2  = "#5b21b6"
PNK   = "#ec4899"
AMB   = "#f59e0b"
GRN   = "#10b981"
RED   = "#ef4444"
ORG   = "#f97316"

T1    = "#f0f2fc"
T2    = "#9aa1c9"
T3    = "#6b73a0"
MONO  = "'Cascadia Code','Consolas','Courier New',monospace"
SANS  = "'Segoe UI','Segoe UI Semibold',Arial,sans-serif"

# Type scale — sized for a full-screen window, not a cramped sidebar.
FS_EYEBROW = 11   # small-caps section labels (DASHBOARD, STORES...)
FS_BODY    = 12   # standard body/paragraph text
FS_SMALL   = 11   # secondary/meta text
FS_TINY    = 10   # least important meta (timestamps etc.)
FS_TITLE   = 21   # page/store titles
FS_STAT    = 34   # big stat card numbers

STATUS_COLOR = {"UPGRADED": GRN, "FAILED": RED, "PENDING": AMB}


def _hx(h):
    h = h.lstrip("#"); return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def rgba(hex_color, alpha):
    r, g, b = _hx(hex_color)
    return f"rgba({r},{g},{b},{alpha})"

def glass_panel_qss(radius=12, base=BG2, border=CY, base_alpha=168, border_alpha=46):
    return (f"background:{rgba(base, base_alpha)};"
            f"border:1px solid {rgba(border, border_alpha)};"
            f"border-radius:{radius}px;")

def glow(widget, color=CY, blur=18, alpha=70, ox=0, oy=6):
    eff = QGraphicsDropShadowEffect(widget)
    eff.setBlurRadius(blur); eff.setOffset(ox, oy)
    r, g, b = _hx(color); eff.setColor(QColor(r, g, b, alpha))
    widget.setGraphicsEffect(eff)
    return eff


# ── Storage ──────────────────────────────────────────────────────────────────
def data_dir():
    appdata = os.environ.get("APPDATA")
    if appdata:
        d = os.path.join(appdata, "Command Centre Pro")
    else:
        d = os.path.join(os.path.expanduser("~"), ".command_centre_pro")
    os.makedirs(d, exist_ok=True)
    return d

def data_path():
    return os.path.join(data_dir(), "UpgradeTracker.json")

def seed_path():
    # Resolve seed_data.json in both frozen (.exe) and dev modes.
    if getattr(sys, "frozen", False):
        base = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    # Look next to this file, then at the bundle/exe root.
    for cand in (os.path.join(base, "seed_data.json"),
                 os.path.join(os.path.dirname(base), "seed_data.json")):
        if os.path.exists(cand):
            return cand
    return os.path.join(base, "seed_data.json")


def _parse_hms(s):
    """Parse an 'H:MM:SS' (or 'MM:SS') duration string into total seconds.
    Returns None if it can't be parsed."""
    if not s:
        return None
    parts = str(s).strip().split(":")
    try:
        parts = [int(p) for p in parts]
    except ValueError:
        return None
    if len(parts) == 3:
        h, m, sec = parts
    elif len(parts) == 2:
        h, m, sec = 0, parts[0], parts[1]
    elif len(parts) == 1:
        h, m, sec = 0, 0, parts[0]
    else:
        return None
    return h * 3600 + m * 60 + sec


# Auto-classify threshold: a 15_SP6 till whose last update is older than this
# is treated as FAILED (it's on the new OS but stopped checking in).
STALE_THRESHOLD_SECONDS = 90 * 60   # 1:30:00


def classify_till(rec):
    """
    Decide a till's upgrade_status from its OS version + last-update time.
      • OS_VER 15_SP6, updated within 1:30:00  -> UPGRADED
      • OS_VER 15_SP6, updated over  1:30:00   -> FAILED (stale / dropped off)
      • OS_VER 12_SP2                          -> PENDING (still on old OS)
      • anything else                          -> unchanged (returns None)
    Returns the new status string, or None to leave the status as-is.
    """
    os_ver = (rec.get("os_ver") or "").strip().upper()
    if os_ver == "15_SP6":
        secs = _parse_hms(rec.get("date_updated"))
        if secs is not None and secs > STALE_THRESHOLD_SECONDS:
            return "FAILED"
        return "UPGRADED"
    if os_ver == "12_SP2":
        return "PENDING"
    return None


class Store:
    __slots__ = ("code", "notes", "tills")
    def __init__(self, code, notes="", tills=None):
        self.code = code
        self.notes = notes
        self.tills = tills or []  # list of dicts

    def counts(self):
        up = sum(1 for t in self.tills if t.get("upgrade_status") == "UPGRADED")
        fa = sum(1 for t in self.tills if t.get("upgrade_status") == "FAILED")
        pe = sum(1 for t in self.tills if t.get("upgrade_status") == "PENDING")
        return up, fa, pe

    def total(self):
        return len(self.tills)

    def pct_complete(self):
        tot = self.total()
        if tot == 0: return 0.0
        up, fa, pe = self.counts()
        return up / tot

    def to_dict(self):
        return {"store": self.code, "notes": self.notes, "tills": self.tills}


class TrackerData(QObject):
    changed = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.stores = {}  # code -> Store
        self.history = []  # list of {"ts":..., "text":...} activity log entries
        self.load()

    # -- persistence --
    def load(self):
        path = data_path()
        if os.path.exists(path):
            try:
                with open(path) as f:
                    raw = json.load(f)
                self.stores = {code: Store(code, s.get("notes",""), s.get("tills", []))
                               for code, s in raw.get("stores", {}).items()}
                self.history = raw.get("history", [])
                return
            except Exception:
                pass
        self.stores = {}
        self.history = []

    def save(self):
        raw = {
            "version": 1,
            "saved": dt.now().isoformat(),
            "stores": {code: s.to_dict() for code, s in self.stores.items()},
            "history": self.history[-500:],
        }
        path = data_path()
        tmp = path + ".tmp"
        with open(tmp, "w") as f:
            json.dump(raw, f, indent=2)
        os.replace(tmp, path)
        self.changed.emit()

    def log(self, text):
        self.history.append({"ts": dt.now().isoformat(), "text": text})

    def has_seed_available(self):
        return os.path.exists(seed_path()) and not os.path.exists(data_path())

    def import_seed(self):
        with open(seed_path()) as f:
            raw = json.load(f)
        self.stores = {code: Store(code, s.get("notes",""), s.get("tills", []))
                       for code, s in raw.get("stores", {}).items()}
        self.log(f"Imported seed data — {len(self.stores)} stores")
        self.save()

    # -- aggregate stats --
    def totals(self):
        tot = up = fa = pe = 0
        for s in self.stores.values():
            tot += s.total()
            u, f, p = s.counts()
            up += u; fa += f; pe += p
        return tot, up, fa, pe

    # -- import from raw SSH-dump text (multi-block CSV) --
    def import_text_dump(self, content, default_status="PENDING"):
        """Returns (stores_touched, tills_added, tills_updated)."""
        # Some copy sources (terminal soft-wrap, certain paste paths) collapse
        # the real line breaks into spaces, turning the whole dump into one
        # line — which then reads as a header with zero data rows and
        # silently imports nothing. Reconstruct real line breaks by splitting
        # before every till-code-looking token (e.g. "no03-cr001"). This is a
        # no-op when proper newlines are already present.
        content = re.sub(r'\s+(?=[a-zA-Z]{2,4}\d{2,3}-cr\d{3,4}\b)', '\n', content)

        blocks = re.split(r'\n\s*\n', content.strip())
        touched = set(); added = 0; updated = 0
        for block in blocks:
            lines = [l for l in block.splitlines() if l.strip()]
            if not lines: continue
            header = [h.strip().upper() for h in lines[0].split(',')]
            if "TILL NUMBER" not in header:
                continue
            idx = {h: i for i, h in enumerate(header)}
            if "UPGRADE STATUS" not in idx and "UPGRADED STATUS" in idx:
                idx["UPGRADE STATUS"] = idx["UPGRADED STATUS"]
            for line in lines[1:]:
                parts = [p.strip() for p in line.split(',')]
                ti = idx.get("TILL NUMBER", 0)
                till_num = parts[ti] if ti < len(parts) else ""
                if not till_num: continue
                store_code = till_num.split('-')[0].upper()
                def g(key):
                    i = idx.get(key)
                    return parts[i].strip() if i is not None and i < len(parts) else ""
                upg = g("UPGRADE STATUS").upper()
                if upg not in ("UPGRADED", "FAILED"):
                    upg = default_status
                rec = {
                    "till": till_num,
                    "last_status": g("LAST STATUS"),
                    "cache_status": g("CACHE STATUS"),
                    "cache_detail": g("CACHE DETAIL"),
                    "date_updated": g("DATE UPDATED"),
                    "pos_version": g("POS VERSION"),
                    "till_ip": g("TILL IP"),
                    "os_ver": g("OS_VER"),
                    "upgrade_status": upg,
                    "failure_type": g("POST/PRE CUTOVER FAILURE") or None,
                    "flag": None,
                }
                # Auto-classify from OS version + update time (overrides the
                # dump's status per the upgrade rules).
                auto = classify_till(rec)
                if auto:
                    rec["upgrade_status"] = auto
                    if auto != "FAILED":
                        rec["failure_type"] = None
                touched.add(store_code)
                if store_code not in self.stores:
                    self.stores[store_code] = Store(store_code)
                store = self.stores[store_code]
                existing = next((t for t in store.tills if t["till"] == till_num), None)
                if existing:
                    existing.update(rec); updated += 1
                else:
                    store.tills.append(rec); added += 1
        self.log(f"Imported dump — {len(touched)} stores, {added} new tills, {updated} refreshed")
        self.save()
        return touched, added, updated

    # -- import from the legacy xlsx workbook format --
    def import_xlsx(self, path):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        skip = {"SUMMARY", "RAW", "UPGRADE REFUSAL LIST"}
        notes_by_store = {}
        if "SUMMARY" in wb.sheetnames:
            ws = wb["SUMMARY"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0] or not isinstance(row[0], str): continue
                code = str(row[0]).strip().upper()
                if not re.match(r'^[A-Za-z]{2}\d{2}$', code): continue
                if len(row) > 7 and row[7]:
                    notes_by_store[code] = str(row[7]).strip()
        touched = set(); added = 0; updated = 0
        for name in wb.sheetnames:
            if name in skip: continue
            ws = wb[name]
            headers = [str(c.value).strip().upper() if c.value else "" for c in ws[1]]
            if "TILL NUMBER" not in headers: continue
            idx = {h: i for i, h in enumerate(headers)}
            if "UPGRADE STATUS" not in idx and "UPGRADED STATUS" in idx:
                idx["UPGRADE STATUS"] = idx["UPGRADED STATUS"]
            code = name.upper()
            if code not in self.stores:
                self.stores[code] = Store(code)
            store = self.stores[code]
            if code in notes_by_store:
                store.notes = notes_by_store[code]
            for row in ws.iter_rows(min_row=2, values_only=True):
                ti = idx.get("TILL NUMBER", 0)
                till_num = row[ti] if ti < len(row) else None
                if not till_num: continue
                till_num = str(till_num).strip()
                def g(key):
                    i = idx.get(key)
                    if i is None or i >= len(row): return ""
                    v = row[i]
                    if v is None: return ""
                    if isinstance(v, datetime.time): return v.strftime("%H:%M:%S")
                    return str(v)
                upg = g("UPGRADE STATUS").upper() or "PENDING"
                if upg not in ("UPGRADED", "FAILED"): upg = "PENDING"
                rec = {
                    "till": till_num, "last_status": g("LAST STATUS"),
                    "cache_status": g("CACHE STATUS"), "cache_detail": g("CACHE DETAIL"),
                    "date_updated": g("DATE UPDATED"), "pos_version": g("POS VERSION"),
                    "till_ip": g("TILL IP"), "os_ver": g("OS_VER"),
                    "upgrade_status": upg, "failure_type": g("POST/PRE CUTOVER FAILURE") or None,
                    "flag": None,
                }
                auto = classify_till(rec)
                if auto:
                    rec["upgrade_status"] = auto
                    if auto != "FAILED":
                        rec["failure_type"] = None
                touched.add(code)
                existing = next((t for t in store.tills if t["till"] == till_num), None)
                if existing:
                    existing.update(rec); updated += 1
                else:
                    store.tills.append(rec); added += 1
        self.log(f"Imported Excel workbook — {len(touched)} stores, {added} new tills, {updated} refreshed")
        self.save()
        return touched, added, updated

    def set_till_status(self, store_code, till_num, status, failure_type=None):
        store = self.stores.get(store_code)
        if not store: return
        for t in store.tills:
            if t["till"] == till_num:
                t["upgrade_status"] = status
                t["failure_type"] = failure_type
                self.log(f"{till_num} -> {status}" + (f" ({failure_type})" if failure_type else ""))
                self.save()
                return

    def set_till_status_bulk(self, store_code, till_nums, status, failure_type=None):
        """Update many tills in one store at once, with a single save — used
        by the multi-select bulk actions so selecting 20 tills doesn't
        trigger 20 separate disk writes."""
        store = self.stores.get(store_code)
        if not store: return 0
        till_set = set(till_nums)
        n = 0
        for t in store.tills:
            if t["till"] in till_set:
                t["upgrade_status"] = status
                t["failure_type"] = failure_type
                n += 1
        if n:
            self.log(f"{n} till(s) in {store_code} -> {status}" + (f" ({failure_type})" if failure_type else ""))
            self.save()
        return n

    def add_store(self, code):
        code = code.strip().upper()
        if code and code not in self.stores:
            self.stores[code] = Store(code)
            self.log(f"Added store {code}")
            self.save()

    def delete_store(self, code):
        if code in self.stores:
            n = len(self.stores[code].tills)
            del self.stores[code]
            self.log(f"Deleted store {code} ({n} tills)")
            self.save()

    def set_store_notes(self, code, notes):
        if code in self.stores:
            self.stores[code].notes = notes
            self.save()


# ── Aurora background ────────────────────────────────────────────────────────
class AuroraBackground(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._t = 0.0
        self._timer = QTimer(self); self._timer.timeout.connect(self._tick); self._timer.start(400)

    def _tick(self):
        self._t += 0.05
        self.update()

    def paintEvent(self, e):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(BG1))
        blobs = [
            (CY,  0.20 + 0.05*math.sin(self._t),       0.15 + 0.04*math.cos(self._t*0.8), 0.55),
            (VIO, 0.80 + 0.05*math.cos(self._t*0.7),   0.28 + 0.05*math.sin(self._t*0.9), 0.60),
            (GRN, 0.35 + 0.06*math.sin(self._t*0.5+1), 0.88 + 0.04*math.cos(self._t*0.6), 0.45),
        ]
        for color, fx, fy, frac in blobs:
            cx, cy = w*fx, h*fy
            rad = max(w, h) * frac
            grad = QRadialGradient(cx, cy, rad)
            r, g, b = _hx(color)
            grad.setColorAt(0.0, QColor(r, g, b, 40))
            grad.setColorAt(0.5, QColor(r, g, b, 14))
            grad.setColorAt(1.0, QColor(r, g, b, 0))
            p.setBrush(QBrush(grad)); p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(QPointF(cx, cy), rad, rad)


def make_label(text, color=T1, size=FS_BODY, bold=False, mono=False):
    l = QLabel(text)
    w = "700" if bold else "400"
    fam = MONO if mono else SANS
    l.setStyleSheet(f"color:{color};font-size:{size}px;font-weight:{w};font-family:{fam};background:transparent;border:none;")
    return l


class Btn(QPushButton):
    def __init__(self, text, color=CY, h=32, filled=False, parent=None):
        super().__init__(text, parent)
        self.setFixedHeight(h)
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        if filled:
            self.setStyleSheet(f"""
                QPushButton{{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 {color},stop:1 {VIO});
                    color:{BG0};border:none;border-radius:{h//2}px;padding:0 16px;font-size:13px;font-weight:700;}}
                QPushButton:hover{{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #33ddff,stop:1 #9256f0);}}
                QPushButton:disabled{{background:{SEP};color:{T3};}}
            """)
        else:
            self.setStyleSheet(f"""
                QPushButton{{background:{rgba(BG3,130)};border:1px solid {color}55;color:{color};
                    border-radius:{h//2}px;padding:0 14px;font-size:12px;font-weight:600;}}
                QPushButton:hover{{background:{color}26;border-color:{color};}}
                QPushButton:pressed{{background:{color}44;}}
                QPushButton:disabled{{background:transparent;border-color:{T3};color:{T3};}}
            """)


class StatCard(QWidget):
    def __init__(self, label, value="—", color=CY, sub="", parent=None):
        super().__init__(parent)
        self._color = color
        self.setObjectName("card")
        self.setStyleSheet(f"QWidget#card{{{glass_panel_qss(radius=14, border=color, base_alpha=150, border_alpha=55)}}}")
        self.setMinimumHeight(96)
        lay = QVBoxLayout(self); lay.setContentsMargins(18,14,18,14); lay.setSpacing(4)
        top = QHBoxLayout(); top.setSpacing(6)
        dot = QLabel("●"); dot.setStyleSheet(f"color:{color};font-size:10px;background:transparent;border:none;"); top.addWidget(dot)
        lbl = make_label(label, T3, FS_SMALL, bold=True)
        lbl.setStyleSheet(lbl.styleSheet() + "letter-spacing:1.5px;")
        top.addWidget(lbl); top.addStretch()
        lay.addLayout(top)
        self._val = QLabel(str(value))
        self._val.setStyleSheet(f"color:{color};font-size:{FS_STAT}px;font-weight:800;font-family:{MONO};background:transparent;border:none;")
        lay.addWidget(self._val)
        if sub:
            lay.addWidget(make_label(sub, T3, FS_SMALL))

    def set_value(self, v):
        self._val.setText(str(v))


class DonutChart(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumSize(150, 150)
        self._segments = []  # list of (value, color)
        self._center_text = ""

    def set_data(self, segments, center_text=""):
        self._segments = segments
        self._center_text = center_text
        self.update()

    def paintEvent(self, e):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        side = min(self.width(), self.height()) - 20
        if side <= 0: return
        rect = QRectF((self.width()-side)/2, (self.height()-side)/2, side, side)
        total = sum(v for v, c in self._segments) or 1
        start = -90 * 16
        thickness = max(14, int(side * 0.16))
        for v, c in self._segments:
            span = int(360 * 16 * (v/total))
            pen = QPen(QColor(c)); pen.setWidth(thickness); pen.setCapStyle(Qt.PenCapStyle.FlatCap)
            p.setPen(pen)
            inset = thickness/2
            p.drawArc(rect.adjusted(inset,inset,-inset,-inset), start, span)
            start += span
        p.setPen(QColor(T1))
        f = p.font(); f.setPointSize(max(10, side//10)); f.setBold(True); p.setFont(f)
        p.drawText(rect, Qt.AlignmentFlag.AlignCenter, self._center_text)


class HBarChart(QWidget):
    """Ranked horizontal bar chart — store code, progress bar, % label."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self._rows = []  # list of (label, pct, color)
        self.setMinimumHeight(40)

    def set_data(self, rows):
        self._rows = rows
        self.setMinimumHeight(max(40, len(rows)*26 + 10))
        self.update()

    def paintEvent(self, e):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w = self.width()
        label_w = 60
        pct_w = 44
        bar_x0 = label_w + 8
        bar_w = max(20, w - label_w - pct_w - 16)
        y = 4
        rh = 20; gap = 6
        f = p.font(); f.setPixelSize(10)
        for label, pct, color in self._rows:
            p.setFont(f)
            p.setPen(QColor(T2))
            p.drawText(QRectF(0, y, label_w, rh), Qt.AlignmentFlag.AlignLeft|Qt.AlignmentFlag.AlignVCenter, label)
            track = QRectF(bar_x0, y+4, bar_w, rh-8)
            p.setPen(Qt.PenStyle.NoPen); p.setBrush(QColor(SEP))
            p.drawRoundedRect(track, 5, 5)
            fill_w = max(3, bar_w * max(0.0, min(1.0, pct)))
            fill = QRectF(bar_x0, y+4, fill_w, rh-8)
            p.setBrush(QColor(color))
            p.drawRoundedRect(fill, 5, 5)
            p.setPen(QColor(T1))
            p.drawText(QRectF(bar_x0+bar_w+8, y, pct_w, rh), Qt.AlignmentFlag.AlignLeft|Qt.AlignmentFlag.AlignVCenter, f"{pct*100:.0f}%")
            y += rh + gap


# ── Title bar ─────────────────────────────────────────────────────────────────
class TitleBar(QFrame):
    def __init__(self, win, on_close, parent=None):
        super().__init__(parent)
        self._win = win; self._drag_pos = None; self._on_close = on_close
        self.setFixedHeight(38)
        self.setStyleSheet(f"QFrame{{background:{rgba(BG0,235)};border-bottom:1px solid {rgba(SEP,140)};}}")
        lay = QHBoxLayout(self); lay.setContentsMargins(14,0,0,0); lay.setSpacing(8)
        icon = QLabel("⬆"); icon.setStyleSheet(f"color:{CY};font-size:15px;background:transparent;border:none;")
        lay.addWidget(icon)
        title = QLabel(f"Upgrade Command Centre  ·  v{APP_VERSION}")
        title.setStyleSheet(f"color:{T2};font-size:12px;font-weight:600;letter-spacing:0.5px;background:transparent;border:none;")
        lay.addWidget(title)
        lay.addStretch()

        def wbtn(text, hover_color, cb):
            b = QPushButton(text); b.setFixedSize(44,38)
            b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            b.setStyleSheet(f"QPushButton{{background:transparent;border:none;color:{T2};font-size:13px;}}QPushButton:hover{{background:{hover_color};color:{T1};}}")
            b.clicked.connect(cb)
            return b

        lay.addWidget(wbtn("─", rgba(BG4,160), win.showMinimized))
        lay.addWidget(wbtn("▢", rgba(BG4,160), self._toggle_max))
        lay.addWidget(wbtn("✕", rgba(RED,200), on_close))

    def _toggle_max(self):
        if self._win.isMaximized(): self._win.showNormal()
        else: self._win.showMaximized()

    def mousePressEvent(self, e):
        if e.button()==Qt.MouseButton.LeftButton:
            self._drag_pos = e.globalPosition().toPoint() - self._win.frameGeometry().topLeft()
        super().mousePressEvent(e)

    def mouseMoveEvent(self, e):
        if self._drag_pos is not None and e.buttons() & Qt.MouseButton.LeftButton:
            if self._win.isMaximized(): self._win.showNormal()
            self._win.move(e.globalPosition().toPoint() - self._drag_pos)
        super().mouseMoveEvent(e)

    def mouseReleaseEvent(self, e):
        self._drag_pos = None
        super().mouseReleaseEvent(e)

    def mouseDoubleClickEvent(self, e):
        self._toggle_max()
        super().mouseDoubleClickEvent(e)


# ── Dashboard panel ───────────────────────────────────────────────────────────
class DashboardPanel(QWidget):
    def __init__(self, data: TrackerData, parent=None):
        super().__init__(parent)
        self.data = data
        self.setStyleSheet("background:transparent;")
        self._build()
        self.refresh()

    def _card(self, title, color):
        w = QWidget(); w.setObjectName("panel")
        w.setStyleSheet(f"QWidget#panel{{{glass_panel_qss(radius=14, border=color, base_alpha=140, border_alpha=40)}}}")
        glow(w, color=color, blur=16, alpha=40, oy=4)
        lay = QVBoxLayout(w); lay.setContentsMargins(18,16,18,16); lay.setSpacing(10)
        lbl = make_label(title, T3, FS_SMALL, bold=True)
        lbl.setStyleSheet(lbl.styleSheet() + "letter-spacing:1.5px;")
        lay.addWidget(lbl)
        return w, lay

    def _build(self):
        outer = QVBoxLayout(self); outer.setContentsMargins(24,20,24,20); outer.setSpacing(16)

        hdr = QHBoxLayout()
        eyebrow = make_label("DASHBOARD", T3, FS_SMALL, bold=True)
        eyebrow.setStyleSheet(eyebrow.styleSheet() + "letter-spacing:2px;")
        hdr.addWidget(eyebrow); hdr.addStretch()
        self._updated_lbl = make_label("", T3, FS_SMALL)
        hdr.addWidget(self._updated_lbl)
        outer.addLayout(hdr)

        # Stat cards
        row = QHBoxLayout(); row.setSpacing(10)
        self._c_total = StatCard("TOTAL TILLS", "—", CY, "across all stores")
        self._c_up    = StatCard("UPGRADED", "—", GRN, "complete")
        self._c_fail  = StatCard("FAILED", "—", RED, "needs retry")
        self._c_pend  = StatCard("PENDING", "—", AMB, "not yet attempted")
        self._c_pct   = StatCard("OVERALL PROGRESS", "—", VIO, "of fleet upgraded")
        for c in [self._c_total, self._c_up, self._c_fail, self._c_pend, self._c_pct]:
            row.addWidget(c)
        outer.addLayout(row)

        mid = QHBoxLayout(); mid.setSpacing(14)

        donut_w, donut_l = self._card("STATUS BREAKDOWN", VIO)
        self._donut = DonutChart(); donut_l.addWidget(self._donut, 1)
        legend = QHBoxLayout(); legend.setSpacing(16)
        for lbl, color in [("Upgraded", GRN), ("Failed", RED), ("Pending", AMB)]:
            dot = QLabel("●"); dot.setStyleSheet(f"color:{color};font-size:12px;background:transparent;border:none;")
            legend.addWidget(dot); legend.addWidget(make_label(lbl, T2, 11))
        legend.addStretch()
        donut_l.addLayout(legend)
        mid.addWidget(donut_w, 1)

        ov_w, ov_l = self._card("STORE STATUS OVERVIEW", CY)
        self._ov_rows = {}
        for key, label, color in [("complete","Complete",GRN), ("attention","Needs attention",AMB), ("failing","Has failures",RED)]:
            row = QHBoxLayout(); row.setSpacing(10)
            dot = QLabel("●"); dot.setStyleSheet(f"color:{color};font-size:13px;background:transparent;border:none;")
            row.addWidget(dot)
            row.addWidget(make_label(label, T2, 12), 1)
            val = make_label("—", color, 20, bold=True, mono=True)
            row.addWidget(val)
            ov_l.addLayout(row)
            self._ov_rows[key] = val
        ov_l.addSpacing(4)
        sep = QFrame(); sep.setFixedHeight(1); sep.setStyleSheet(f"background:{SEP};border:none;")
        ov_l.addWidget(sep)
        ov_l.addSpacing(4)

        # Pre/Post cutover failure breakdown (from each till's failure_type)
        for key, label, color in [("pre","Pre-cutover failures",AMB),
                                    ("post","Post-cutover failures",RED)]:
            crow = QHBoxLayout(); crow.setSpacing(10)
            dot = QLabel("◆"); dot.setStyleSheet(f"color:{color};font-size:11px;background:transparent;border:none;")
            crow.addWidget(dot)
            crow.addWidget(make_label(label, T2, 12), 1)
            cval = make_label("—", color, 20, bold=True, mono=True)
            crow.addWidget(cval)
            ov_l.addLayout(crow)
            self._ov_rows[key] = cval
        ov_l.addSpacing(4)
        sep2 = QFrame(); sep2.setFixedHeight(1); sep2.setStyleSheet(f"background:{SEP};border:none;")
        ov_l.addWidget(sep2)
        ov_l.addSpacing(4)
        tot_row = QHBoxLayout()
        tot_row.addWidget(make_label("Total stores tracked", T3, 11), 1)
        self._ov_total = make_label("—", T1, 20, bold=True, mono=True)
        tot_row.addWidget(self._ov_total)
        ov_l.addLayout(tot_row)
        ov_l.addStretch()
        mid.addWidget(ov_w, 1)
        outer.addLayout(mid, 1)

        prog_w, prog_l = self._card("STORE PROGRESS — NEEDS ATTENTION FIRST", AMB)
        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet(f"QScrollArea{{background:transparent;border:none;}}QScrollBar:vertical{{background:transparent;width:5px;}}QScrollBar::handle:vertical{{background:{SEP};border-radius:2px;}}")
        inner = QWidget(); inner.setStyleSheet("background:transparent;")
        self._store_bar_layout = QVBoxLayout(inner); self._store_bar_layout.setContentsMargins(0,0,0,0)
        scroll.setWidget(inner)
        prog_l.addWidget(scroll, 1)
        outer.addWidget(prog_w, 2)

    def refresh(self):
        tot, up, fa, pe = self.data.totals()
        self._c_total.set_value(tot)
        self._c_up.set_value(up)
        self._c_fail.set_value(fa)
        self._c_pend.set_value(pe)
        pct = (up/tot*100) if tot else 0
        self._c_pct.set_value(f"{pct:.0f}%")
        self._donut.set_data([(up, GRN), (fa, RED), (pe, AMB)], f"{tot}")

        complete = attention = failing = 0
        pre = post = 0
        for s in self.data.stores.values():
            u, f, p = s.counts()
            if f > 0: failing += 1
            elif p > 0: attention += 1
            else: complete += 1
            for t in s.tills:
                ft = (t.get("failure_type") or "").upper()
                if ft == "PRE.CUTOVER": pre += 1
                elif ft == "POST.CUTOVER": post += 1
        self._ov_rows["complete"].setText(str(complete))
        self._ov_rows["attention"].setText(str(attention))
        self._ov_rows["failing"].setText(str(failing))
        self._ov_rows["pre"].setText(str(pre))
        self._ov_rows["post"].setText(str(post))
        self._ov_total.setText(str(len(self.data.stores)))

        # Clear and rebuild the per-store bar chart, worst-first (most pending+failed)
        while self._store_bar_layout.count():
            item = self._store_bar_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
        ranked = sorted(self.data.stores.values(), key=lambda s: (-(s.counts()[1]+s.counts()[2]), s.code))
        rows = []
        for s in ranked:
            u, f, p = s.counts()
            t = s.total() or 1
            rows.append((s.code, u/t, GRN if p==0 and f==0 else (RED if f>0 else AMB)))
        chart = HBarChart(); chart.set_data(rows)
        self._store_bar_layout.addWidget(chart)

        self._updated_lbl.setText("Updated " + dt.now().strftime("%H:%M:%S"))


# ── Store list card ───────────────────────────────────────────────────────────
class StoreListCard(QFrame):
    clicked = pyqtSignal(str)
    def __init__(self, store: Store, parent=None):
        super().__init__(parent)
        self.code = store.code
        self.setFixedHeight(72)
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        u, f, p = store.counts()
        tot = store.total() or 1
        pct = u / tot
        color = GRN if (p==0 and f==0) else (RED if f>0 else AMB)
        self.setStyleSheet(f"StoreListCard{{background:{rgba(BG2,120)};border:1px solid {rgba(SEP,90)};border-radius:12px;}}"
                            f"StoreListCard:hover{{background:{rgba(BG4,150)};border-color:{rgba(CY,90)};}}")
        lay = QHBoxLayout(self); lay.setContentsMargins(16,8,16,8); lay.setSpacing(12)
        bar = QFrame(); bar.setFixedWidth(4); bar.setStyleSheet(f"background:{color};border-radius:2px;")
        lay.addWidget(bar)
        col = QVBoxLayout(); col.setSpacing(4)
        top = QHBoxLayout()
        top.addWidget(make_label(store.code, T1, 16, bold=True))
        top.addStretch()
        top.addWidget(make_label(f"{u}/{tot} upgraded", T2, 9))
        col.addLayout(top)
        track = QWidget(); track.setFixedHeight(6)
        track.setStyleSheet(f"background:{SEP};border-radius:3px;")
        fill = QWidget(track); fill.setStyleSheet(f"background:{color};border-radius:3px;")
        fill.setGeometry(0, 0, int(300*pct), 6)
        self._track = track; self._fill = fill; self._pct = pct
        col.addWidget(track)
        lay.addLayout(col, 1)
        badges = QHBoxLayout(); badges.setSpacing(6)
        if f: badges.addWidget(self._pill(f"{f} failed", RED))
        if p: badges.addWidget(self._pill(f"{p} pending", AMB))
        if not f and not p: badges.addWidget(self._pill("complete", GRN))
        lay.addLayout(badges)

    def _pill(self, text, color):
        l = QLabel(text)
        l.setStyleSheet(f"color:{color};background:{rgba(color,32)};border:1px solid {color}55;border-radius:8px;padding:2px 8px;font-size:11px;font-weight:700;")
        return l

    def resizeEvent(self, e):
        super().resizeEvent(e)
        self._fill.setGeometry(0, 0, int(self._track.width()*self._pct), 6)

    def mousePressEvent(self, e):
        if e.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit(self.code)
        super().mousePressEvent(e)


# ── Till row (editable status) ────────────────────────────────────────────────
class TillRow(QFrame):
    changed = pyqtSignal(str, str, object)  # till, status, failure_type
    selection_changed = pyqtSignal(str, bool)  # till, is_selected
    def __init__(self, till_rec, parent=None):
        super().__init__(parent)
        self.till = till_rec
        self.setFixedHeight(46)
        self.setStyleSheet(f"TillRow{{background:{rgba(BG2,100)};border-radius:8px;}}TillRow:hover{{background:{rgba(BG4,130)};}}")
        lay = QHBoxLayout(self); lay.setContentsMargins(12,0,12,0); lay.setSpacing(10)

        self._check = QCheckBox()
        self._check.setStyleSheet(f"QCheckBox::indicator{{width:16px;height:16px;border-radius:4px;border:1px solid {SEP};background:{rgba(BG0,150)};}}QCheckBox::indicator:checked{{background:{CY};border-color:{CY};}}")
        self._check.toggled.connect(lambda checked: self.selection_changed.emit(self.till["till"], checked))
        lay.addWidget(self._check, 0)

        lay.addWidget(make_label(till_rec["till"], CY, 10, mono=True, bold=True), 0)
        lay.addWidget(make_label(till_rec.get("last_status","") or "—", T2, 9), 1)
        lay.addWidget(make_label(till_rec.get("till_ip","") or "—", T3, FS_SMALL, mono=True), 0)
        lay.addWidget(make_label(till_rec.get("pos_version","") or "—", T3, FS_SMALL, mono=True), 0)

        # OS version — colour-coded: 15_SP6 (new) green-ish, 12_SP2 (old) amber
        osv = (till_rec.get("os_ver","") or "").strip() or "—"
        os_color = GRN if osv.upper() == "15_SP6" else (AMB if osv.upper() == "12_SP2" else T3)
        os_lbl = make_label(osv, os_color, FS_SMALL, mono=True, bold=True)
        os_lbl.setFixedWidth(58); os_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(os_lbl, 0)

        # Uptime / time since last update (H:MM:SS). Amber if stale (>1:30:00).
        upt = (till_rec.get("date_updated","") or "").strip() or "—"
        upt_secs = _parse_hms(upt)
        upt_color = AMB if (upt_secs is not None and upt_secs > STALE_THRESHOLD_SECONDS) else T3
        upt_lbl = make_label(upt, upt_color, FS_SMALL, mono=True)
        upt_lbl.setFixedWidth(62); upt_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(upt_lbl, 0)

        self._status_btn = QToolButton()
        self._status_btn.setText(till_rec.get("upgrade_status","PENDING"))
        self._status_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self._status_btn.setFixedWidth(125); self._status_btn.setFixedHeight(28)
        self._status_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        status_menu = QMenu(self._status_btn)
        status_menu.setStyleSheet(f"QMenu{{background:{BG2};border:1px solid {SEP};border-radius:8px;padding:4px;}}QMenu::item{{color:{T1};padding:6px 16px;border-radius:8px;font-size:11px;}}QMenu::item:selected{{background:{rgba(CY,40)};color:{CY};}}")
        for opt in ["PENDING", "UPGRADED", "FAILED"]:
            act = status_menu.addAction(opt)
            act.triggered.connect(lambda checked=False, o=opt: self._set_status(o))
        self._status_btn.setMenu(status_menu)
        color = STATUS_COLOR.get(till_rec.get("upgrade_status","PENDING"), AMB)
        self._style_status_btn(color)
        lay.addWidget(self._status_btn, 0)

        self._fail_btn = QToolButton()
        ft0 = till_rec.get("failure_type") or "—"
        if ft0 not in ("PRE.CUTOVER","POST.CUTOVER"): ft0 = "—"
        self._fail_btn.setText(ft0)
        self._fail_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self._fail_btn.setFixedWidth(122); self._fail_btn.setFixedHeight(28)
        self._fail_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._fail_btn.setStyleSheet(f"QToolButton{{background:{rgba(BG0,140)};border:1px solid {SEP};border-radius:8px;color:{T2};font-size:11px;padding:2px 6px;}}QToolButton::menu-indicator{{width:0px;}}")
        fail_menu = QMenu(self._fail_btn)
        fail_menu.setStyleSheet(f"QMenu{{background:{BG2};border:1px solid {SEP};border-radius:8px;padding:4px;}}QMenu::item{{color:{T1};padding:6px 16px;border-radius:8px;font-size:11px;}}QMenu::item:selected{{background:{rgba(CY,40)};color:{CY};}}")
        for opt in ["—", "PRE.CUTOVER", "POST.CUTOVER"]:
            act = fail_menu.addAction(opt)
            act.triggered.connect(lambda checked=False, o=opt: self._set_fail_type(o))
        self._fail_btn.setMenu(fail_menu)
        self._fail_btn.setVisible(till_rec.get("upgrade_status")=="FAILED")
        lay.addWidget(self._fail_btn, 0)

    def set_checked(self, checked):
        self._check.blockSignals(True)
        self._check.setChecked(checked)
        self._check.blockSignals(False)

    def _style_status_btn(self, color):
        self._status_btn.setStyleSheet(f"QToolButton{{background:{rgba(color,32)};border:1px solid {color}66;border-radius:8px;color:{color};font-size:11px;font-weight:700;padding:2px 8px;}}QToolButton::menu-indicator{{width:0px;}}")

    def _set_status(self, text):
        self._status_btn.setText(text)
        color = STATUS_COLOR.get(text, AMB)
        self._style_status_btn(color)
        self._fail_btn.setVisible(text=="FAILED")
        ft = self._fail_btn.text() if text=="FAILED" and self._fail_btn.text()!="—" else None
        self.changed.emit(self.till["till"], text, ft)

    def _set_fail_type(self, text):
        self._fail_btn.setText(text)
        if self._status_btn.text()=="FAILED":
            self.changed.emit(self.till["till"], "FAILED", None if text=="—" else text)


# ── Store detail panel ────────────────────────────────────────────────────────
class StoreDetailPanel(QWidget):
    back = pyqtSignal()
    def __init__(self, data: TrackerData, parent=None):
        super().__init__(parent)
        self.data = data; self.code = None
        self._rows = {}       # till -> TillRow
        self._selected = set()  # till numbers currently checked
        self.setStyleSheet("background:transparent;")
        self._build()

    def _build(self):
        root = QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(10)
        hdr = QHBoxLayout()
        back_btn = Btn("← All stores", T2, h=28); back_btn.clicked.connect(self.back.emit)
        hdr.addWidget(back_btn)
        hdr.addStretch()
        del_btn = Btn("🗑  Delete store", RED, h=28); del_btn.clicked.connect(self._on_delete_store)
        hdr.addWidget(del_btn)
        root.addLayout(hdr)
        self._title = make_label("", T1, FS_TITLE, bold=True)
        root.addWidget(self._title)

        stats = QHBoxLayout(); stats.setSpacing(8)
        self._s_up = StatCard("UPGRADED", "—", GRN); self._s_fa = StatCard("FAILED", "—", RED)
        self._s_pe = StatCard("PENDING", "—", AMB); self._s_tot = StatCard("TOTAL", "—", CY)
        for c in [self._s_up, self._s_fa, self._s_pe, self._s_tot]:
            c.setMinimumHeight(70); stats.addWidget(c)
        root.addLayout(stats)

        notes_row = QHBoxLayout()
        notes_row.addWidget(make_label("Store notes:", T3, FS_SMALL))
        self._notes_edit = QLineEdit()
        self._notes_edit.setStyleSheet(f"QLineEdit{{background:{rgba(BG0,150)};border:1px solid {SEP};border-radius:8px;color:{T1};font-size:12px;padding:4px 10px;}}")
        self._notes_edit.editingFinished.connect(self._on_notes_edit)
        notes_row.addWidget(self._notes_edit, 1)
        root.addLayout(notes_row)

        # -- Bulk-action toolbar --
        bar = QWidget(); bar.setObjectName("bulkbar")
        bar.setStyleSheet(f"QWidget#bulkbar{{{glass_panel_qss(radius=10, base=BG2, border=CY, base_alpha=140, border_alpha=45)}}}")
        bl = QHBoxLayout(bar); bl.setContentsMargins(12,8,12,8); bl.setSpacing(10)
        self._select_all_cb = QCheckBox("Select all")
        self._select_all_cb.setStyleSheet(f"QCheckBox{{color:{T2};font-size:12px;}}QCheckBox::indicator{{width:16px;height:16px;border-radius:4px;border:1px solid {SEP};background:{rgba(BG0,150)};}}QCheckBox::indicator:checked{{background:{CY};border-color:{CY};}}")
        self._select_all_cb.toggled.connect(self._on_select_all)
        bl.addWidget(self._select_all_cb)
        self._selected_lbl = make_label("0 selected", T3, FS_SMALL)
        bl.addWidget(self._selected_lbl)
        bl.addStretch()
        bl.addWidget(make_label("Mark selected as:", T3, FS_SMALL))
        up_btn = Btn("Upgraded", GRN, h=28); up_btn.clicked.connect(lambda: self._bulk_set("UPGRADED"))
        fa_btn = Btn("Failed", RED, h=28); fa_btn.clicked.connect(lambda: self._bulk_set("FAILED"))
        pe_btn = Btn("Pending", AMB, h=28); pe_btn.clicked.connect(lambda: self._bulk_set("PENDING"))
        for b in [up_btn, fa_btn, pe_btn]: bl.addWidget(b)

        # separator between status actions and the copy action
        sep = QLabel(); sep.setFixedSize(1, 20)
        sep.setStyleSheet(f"background:{SEP};")
        bl.addWidget(sep)

        self._copy_ips_btn = Btn("⧉  Copy IPs", CY, h=28)
        self._copy_ips_btn.clicked.connect(self._copy_selected_ips)
        bl.addWidget(self._copy_ips_btn)
        root.addWidget(bar)

        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet(f"QScrollArea{{background:transparent;border:none;}}QScrollBar:vertical{{background:transparent;width:5px;}}QScrollBar::handle:vertical{{background:{SEP};border-radius:2px;}}")
        inner = QWidget(); inner.setStyleSheet("background:transparent;")
        self._till_layout = QVBoxLayout(inner); self._till_layout.setSpacing(4)
        scroll.setWidget(inner)
        root.addWidget(scroll, 1)

    def load(self, code):
        self.code = code
        store = self.data.stores.get(code)
        if not store: return
        self._title.setText(code)
        u, f, p = store.counts()
        self._s_up.set_value(u); self._s_fa.set_value(f); self._s_pe.set_value(p); self._s_tot.set_value(store.total())
        self._notes_edit.setText(store.notes)
        self._selected = set()
        self._rows = {}
        self._select_all_cb.blockSignals(True); self._select_all_cb.setChecked(False); self._select_all_cb.blockSignals(False)
        self._update_selected_label()
        while self._till_layout.count():
            item = self._till_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
        for t in sorted(store.tills, key=lambda x: x["till"]):
            row = TillRow(t)
            row.changed.connect(self._on_till_changed)
            row.selection_changed.connect(self._on_row_selection_changed)
            self._till_layout.addWidget(row)
            self._rows[t["till"]] = row
        self._till_layout.addStretch()

    def _copy_selected_ips(self):
        if not self._selected or not self.code:
            self._flash_copy_btn("Select tills first", RED)
            return
        store = self.data.stores.get(self.code)
        if not store:
            return
        # Collect IPs for the selected tills, preserving till order, skipping blanks.
        ips = []
        for t in sorted(store.tills, key=lambda x: x["till"]):
            if t["till"] in self._selected:
                ip = (t.get("till_ip") or "").strip()
                if ip and ip != "—":
                    ips.append(ip)
        if not ips:
            self._flash_copy_btn("No IPs found", RED)
            return
        from PyQt6.QtWidgets import QApplication
        QApplication.clipboard().setText("\n".join(ips))
        self._flash_copy_btn(f"✓ Copied {len(ips)} IP{'s' if len(ips) != 1 else ''}", GRN)

    def _flash_copy_btn(self, msg, color):
        """Briefly recolor + relabel the copy button as feedback, then restore."""
        h = 28
        self._copy_ips_btn.setText(msg)
        self._copy_ips_btn.setStyleSheet(f"""
            QPushButton{{background:{color}26;border:1px solid {color};color:{color};
                border-radius:{h//2}px;padding:0 14px;font-size:12px;font-weight:600;}}
        """)
        def _restore():
            self._copy_ips_btn.setText("⧉  Copy IPs")
            self._copy_ips_btn.setStyleSheet(f"""
                QPushButton{{background:{rgba(BG3,130)};border:1px solid {CY}55;color:{CY};
                    border-radius:{h//2}px;padding:0 14px;font-size:12px;font-weight:600;}}
                QPushButton:hover{{background:{CY}26;border-color:{CY};}}
                QPushButton:pressed{{background:{CY}44;}}
            """)
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(1400, _restore)

    def _on_row_selection_changed(self, till, checked):
        if checked: self._selected.add(till)
        else: self._selected.discard(till)
        self._update_selected_label()

    def _update_selected_label(self):
        n = len(self._selected)
        self._selected_lbl.setText(f"{n} selected" if n else "0 selected")

    def _on_select_all(self, checked):
        for till, row in self._rows.items():
            row.set_checked(checked)
            if checked: self._selected.add(till)
            else: self._selected.discard(till)
        self._update_selected_label()

    def _bulk_set(self, status):
        if not self._selected or not self.code:
            return
        self.data.set_till_status_bulk(self.code, list(self._selected), status, None)
        self.load(self.code)  # reload to reflect new statuses and reset selection

    def _on_delete_store(self):
        if not self.code: return
        box = QMessageBox(self)
        box.setWindowTitle("Delete store")
        box.setText(f"Delete {self.code} and all {self.data.stores[self.code].total()} tracked tills? This can't be undone.")
        box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        box.setDefaultButton(QMessageBox.StandardButton.No)
        if box.exec() == QMessageBox.StandardButton.Yes:
            self.data.delete_store(self.code)
            self.back.emit()

    def _on_till_changed(self, till, status, ftype):
        self.data.set_till_status(self.code, till, status, ftype)
        # refresh the stat cards inline without rebuilding the whole till list
        store = self.data.stores.get(self.code)
        if store:
            u, f, p = store.counts()
            self._s_up.set_value(u); self._s_fa.set_value(f); self._s_pe.set_value(p)

    def _on_notes_edit(self):
        if self.code:
            self.data.set_store_notes(self.code, self._notes_edit.text())


# ── Stores panel (list + detail, swappable) ──────────────────────────────────
class StoresPanel(QWidget):
    def __init__(self, data: TrackerData, parent=None):
        super().__init__(parent)
        self.data = data
        self.setStyleSheet("background:transparent;")
        self._build()
        self.refresh_list()

    def _build(self):
        root = QVBoxLayout(self); root.setContentsMargins(24,20,24,20); root.setSpacing(14)
        self._stack = QStackedWidget()

        # -- list page --
        list_page = QWidget(); list_page.setStyleSheet("background:transparent;")
        ll = QVBoxLayout(list_page); ll.setContentsMargins(0,0,0,0); ll.setSpacing(10)

        hdr = QHBoxLayout()
        eyebrow = make_label("STORES", T3, FS_SMALL, bold=True); eyebrow.setStyleSheet(eyebrow.styleSheet()+"letter-spacing:2px;")
        hdr.addWidget(eyebrow); hdr.addStretch()
        add_btn = Btn("+ Add store", CY, h=28); add_btn.clicked.connect(self._add_store)
        hdr.addWidget(add_btn)
        ll.addLayout(hdr)

        search_row = QHBoxLayout()
        self._search = QLineEdit(); self._search.setPlaceholderText("Search stores…")
        self._search.setFixedHeight(38)
        self._search.setStyleSheet(f"QLineEdit{{background:{rgba(BG2,140)};border:1px solid {SEP};border-radius:12px;color:{T1};font-size:13px;padding:0 12px;}}")
        self._search.textChanged.connect(lambda _: self.refresh_list())
        search_row.addWidget(self._search)
        self._filter_combo = QComboBox(); self._filter_combo.addItems(["All", "Incomplete", "Complete", "Has failures"])
        self._filter_combo.setFixedHeight(38); self._filter_combo.setFixedWidth(150)
        self._filter_combo.setStyleSheet(f"QComboBox{{background:{rgba(BG2,140)};border:1px solid {SEP};border-radius:12px;color:{T2};font-size:12px;padding:0 10px;}}")
        self._filter_combo.currentTextChanged.connect(lambda _: self.refresh_list())
        search_row.addWidget(self._filter_combo)
        ll.addLayout(search_row)

        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet(f"QScrollArea{{background:transparent;border:none;}}QScrollBar:vertical{{background:transparent;width:5px;}}QScrollBar::handle:vertical{{background:{SEP};border-radius:2px;}}")
        inner = QWidget(); inner.setStyleSheet("background:transparent;")
        self._list_layout = QVBoxLayout(inner); self._list_layout.setSpacing(6)
        scroll.setWidget(inner)
        ll.addWidget(scroll, 1)

        self._stack.addWidget(list_page)

        # -- detail page --
        self._detail = StoreDetailPanel(self.data)
        self._detail.back.connect(self._on_back_to_list)
        self._stack.addWidget(self._detail)

        root.addWidget(self._stack)

    def refresh_list(self):
        while self._list_layout.count():
            item = self._list_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
        q = self._search.text().strip().lower()
        f = self._filter_combo.currentText()
        stores = sorted(self.data.stores.values(), key=lambda s: s.code)
        for s in stores:
            if q and q not in s.code.lower(): continue
            u, fa, p = s.counts()
            if f == "Incomplete" and p==0 and fa==0: continue
            if f == "Complete" and (p>0 or fa>0): continue
            if f == "Has failures" and fa==0: continue
            card = StoreListCard(s)
            card.clicked.connect(self._open_store)
            self._list_layout.addWidget(card)
        self._list_layout.addStretch()

    def _open_store(self, code):
        self._detail.load(code)
        self._stack.setCurrentIndex(1)

    def _on_back_to_list(self):
        self._stack.setCurrentIndex(0)
        self.refresh_list()

    def _add_store(self):
        code, ok = QInputDialog.getText(self, "Add store", "Store code (e.g. GC48):")
        if ok and code.strip():
            self.data.add_store(code.strip().upper())
            self.refresh_list()


# ── Import panel ──────────────────────────────────────────────────────────────
class ImportPanel(QWidget):
    imported = pyqtSignal()
    def __init__(self, data: TrackerData, parent=None):
        super().__init__(parent)
        self.data = data
        self.setStyleSheet("background:transparent;")
        self._build()

    def _build(self):
        root = QVBoxLayout(self); root.setContentsMargins(24,20,24,20); root.setSpacing(16)
        eyebrow = make_label("IMPORT NEW DATA", T3, FS_SMALL, bold=True); eyebrow.setStyleSheet(eyebrow.styleSheet()+"letter-spacing:2px;")
        root.addWidget(eyebrow)

        desc = make_label("Paste a raw cache-health dump (the same CSV block format you get from an SSH "
                           "session) below. New tills become PENDING; tills that already exist get refreshed "
                           "in place — statuses you've already set are kept unless the dump itself says otherwise.",
                           T2, 10)
        desc.setWordWrap(True)
        root.addWidget(desc)

        paste_card = QWidget(); paste_card.setObjectName("panel")
        paste_card.setStyleSheet(f"QWidget#panel{{{glass_panel_qss(radius=14, border=CY, base_alpha=140, border_alpha=40)}}}")
        pl = QVBoxLayout(paste_card); pl.setContentsMargins(18,16,18,16); pl.setSpacing(12)
        self._paste = QTextEdit()
        self._paste.setPlaceholderText("TILL NUMBER,LAST STATUS,CACHE STATUS,CACHE DETAIL,DATE UPDATED,POS VERSION,TILL IP,OS_VER,...\ngc48-cr001,LOGON SCREEN,ACTIVE,OK,00:09:17,26.2_002,10.192.134.11,12_SP2,...")
        self._paste.setStyleSheet(f"QTextEdit{{background:{rgba(BG0,150)};border:1px solid {SEP};border-radius:8px;color:{T1};font-size:12px;font-family:{MONO};padding:8px;}}")
        self._paste.setMinimumHeight(260)
        pl.addWidget(self._paste)

        row = QHBoxLayout(); row.setSpacing(10)
        import_btn = Btn("Import pasted data", CY, h=34, filled=True)
        import_btn.clicked.connect(self._do_import_text)
        row.addWidget(import_btn)
        clear_btn = Btn("Clear", T3, h=34); clear_btn.clicked.connect(self._paste.clear)
        row.addWidget(clear_btn)
        row.addStretch()
        xlsx_btn = Btn("Import legacy Excel workbook…", VIO, h=34)
        xlsx_btn.clicked.connect(self._do_import_xlsx)
        row.addWidget(xlsx_btn)
        pl.addLayout(row)

        self._result_lbl = make_label("", GRN, 10)
        self._result_lbl.setWordWrap(True)
        pl.addWidget(self._result_lbl)

        root.addWidget(paste_card)

        # Recent activity log
        log_card = QWidget(); log_card.setObjectName("panel")
        log_card.setStyleSheet(f"QWidget#panel{{{glass_panel_qss(radius=14, border=T3, base_alpha=120, border_alpha=30)}}}")
        ll = QVBoxLayout(log_card); ll.setContentsMargins(18,16,18,16); ll.setSpacing(10)
        ll.addWidget(make_label("RECENT ACTIVITY", T3, FS_SMALL, bold=True))
        self._log_area = QTextEdit(); self._log_area.setReadOnly(True)
        self._log_area.setStyleSheet(f"QTextEdit{{background:{rgba(BG0,140)};border:none;color:{T2};font-size:11px;font-family:{MONO};}}")
        self._log_area.setMaximumHeight(160)
        ll.addWidget(self._log_area)
        root.addWidget(log_card)
        root.addStretch()

        self.refresh_log()

    def _do_import_text(self):
        text = self._paste.toPlainText().strip()
        if not text:
            self._result_lbl.setStyleSheet(self._result_lbl.styleSheet().replace(GRN, RED))
            self._result_lbl.setText("Paste some data first.")
            return
        try:
            touched, added, updated = self.data.import_text_dump(text)
            self._result_lbl.setStyleSheet(f"color:{GRN};font-size:12px;background:transparent;border:none;")
            self._result_lbl.setText(f"Imported {len(touched)} store(s): {', '.join(sorted(touched))} — {added} new till(s), {updated} refreshed.")
            self._paste.clear()
            self.refresh_log()
            self.imported.emit()
        except Exception as e:
            self._result_lbl.setStyleSheet(f"color:{RED};font-size:12px;background:transparent;border:none;")
            self._result_lbl.setText(f"Import failed: {e}")

    def _do_import_xlsx(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import legacy Excel workbook", "", "Excel files (*.xlsx)")
        if not path: return
        try:
            touched, added, updated = self.data.import_xlsx(path)
            self._result_lbl.setStyleSheet(f"color:{GRN};font-size:12px;background:transparent;border:none;")
            self._result_lbl.setText(f"Imported {len(touched)} store(s) from workbook — {added} new till(s), {updated} refreshed.")
            self.refresh_log()
            self.imported.emit()
        except Exception as e:
            self._result_lbl.setStyleSheet(f"color:{RED};font-size:12px;background:transparent;border:none;")
            self._result_lbl.setText(f"Import failed: {e}")

    def refresh_log(self):
        lines = []
        for entry in reversed(self.data.history[-30:]):
            ts = entry.get("ts","")
            try:
                ts = dt.fromisoformat(ts).strftime("%d %b %H:%M")
            except Exception:
                pass
            lines.append(f"[{ts}] {entry.get('text','')}")
        self._log_area.setPlainText("\n".join(lines) if lines else "No activity yet.")


# ── Share panel ────────────────────────────────────────────────────────────────
# ── Shareable report image ────────────────────────────────────────────────────
def render_report_image(data: TrackerData) -> QPixmap:
    """Paints a standalone shareable summary image directly — deliberately NOT
    a screenshot of the live dashboard. Grabbing the real dashboard widget in
    isolation loses the aurora background it depends on for its glass look
    (falls back to a flat gray) and only captures whatever's currently
    scrolled into view. This instead sizes itself to fit every store, always,
    and is designed to be read as a static image rather than an interactive
    screen."""
    W = 1120
    stores = sorted(data.stores.values(), key=lambda s: (-(s.counts()[1]+s.counts()[2]), s.code))
    n = len(stores)
    HEADER_H = 200
    ROW_H = 30
    TABLE_HDR_H = 34
    FOOTER_H = 44
    PAD = 32
    table_h = TABLE_HDR_H + n*ROW_H
    H = HEADER_H + table_h + FOOTER_H + PAD

    pix = QPixmap(W, H)
    pix.fill(QColor(BG1))
    p = QPainter(pix)
    p.setRenderHint(QPainter.RenderHint.Antialiasing)
    p.setRenderHint(QPainter.RenderHint.TextAntialiasing)

    # Subtle static glow behind the header only — keeps the aurora flavor
    # without risking legibility over the data table below.
    for color, fx, fy, frac in [(CY, 0.12, 0.05, 0.55), (VIO, 0.92, 0.15, 0.55)]:
        cx, cy = W*fx, HEADER_H*fy + 20
        rad = W * frac
        grad = QRadialGradient(cx, cy, rad)
        r, g, b = _hx(color)
        grad.setColorAt(0.0, QColor(r,g,b,50)); grad.setColorAt(0.6, QColor(r,g,b,12)); grad.setColorAt(1.0, QColor(r,g,b,0))
        p.setBrush(QBrush(grad)); p.setPen(Qt.PenStyle.NoPen)
        p.drawEllipse(QPointF(cx,cy), rad, rad)

    tot, up, fa, pe = data.totals()
    pct = (up/tot*100) if tot else 0

    # -- Header: title + timestamp --
    p.setPen(QColor(T1))
    f = QFont(SANS.split(",")[0].strip("'\""), 24, QFont.Weight.Bold)
    p.setFont(f)
    p.drawText(QRectF(PAD, 20, W-2*PAD, 34), Qt.AlignmentFlag.AlignLeft, "POS Upgrade Progress")
    p.setPen(QColor(T3))
    f2 = QFont(SANS.split(",")[0].strip("'\""), 11)
    p.setFont(f2)
    p.drawText(QRectF(PAD, 54, W-2*PAD, 20), Qt.AlignmentFlag.AlignLeft,
               f"Generated {dt.now().strftime('%d %b %Y, %H:%M')}  ·  {n} stores tracked")

    # -- Stat blocks --
    stats = [("TOTAL TILLS", str(tot), CY), ("UPGRADED", str(up), GRN),
             ("FAILED", str(fa), RED), ("PENDING", str(pe), AMB),
             ("PROGRESS", f"{pct:.0f}%", VIO)]
    block_w = (W - 2*PAD) / len(stats)
    for i, (label, val, color) in enumerate(stats):
        x = PAD + i*block_w
        p.setPen(QColor(T3))
        fL = QFont(SANS.split(",")[0].strip("'\""), 10, QFont.Weight.Bold)
        p.setFont(fL)
        p.drawText(QRectF(x, 100, block_w-10, 16), Qt.AlignmentFlag.AlignLeft, label)
        p.setPen(QColor(color))
        fV = QFont(MONO.split(",")[0].strip("'\""), 26, QFont.Weight.Bold)
        p.setFont(fV)
        p.drawText(QRectF(x, 118, block_w-10, 40), Qt.AlignmentFlag.AlignLeft, val)

    # -- Divider --
    p.setPen(QPen(QColor(SEP), 1))
    p.drawLine(QPointF(PAD, HEADER_H-6), QPointF(W-PAD, HEADER_H-6))

    # -- Table header --
    ty = HEADER_H
    cols = [("STORE", 90), ("TOTAL", 60), ("UP", 55), ("FAIL", 55), ("PEND", 55), ("PROGRESS", 380), ("%", 50)]
    p.setPen(QColor(T3))
    fH = QFont(SANS.split(",")[0].strip("'\""), 10, QFont.Weight.Bold)
    p.setFont(fH)
    cx = PAD
    col_x = []
    for label, w in cols:
        col_x.append(cx)
        p.drawText(QRectF(cx, ty+8, w, 18), Qt.AlignmentFlag.AlignLeft, label)
        cx += w
    ty += TABLE_HDR_H
    p.setPen(QPen(QColor(SEP), 1))
    p.drawLine(QPointF(PAD, ty), QPointF(W-PAD, ty))

    # -- Table rows --
    fRow = QFont(SANS.split(",")[0].strip("'\""), 11)
    fRowBold = QFont(SANS.split(",")[0].strip("'\""), 12, QFont.Weight.Bold)
    fMono = QFont(MONO.split(",")[0].strip("'\""), 11)
    for i, s in enumerate(stores):
        u, f, pd = s.counts(); t = s.total() or 1
        row_pct = u / t
        color = GRN if (f==0 and pd==0) else (RED if f>0 else AMB)
        ry = ty + i*ROW_H
        if i % 2 == 0:
            r,g,b = _hx(BG2)
            p.fillRect(QRectF(PAD-6, ry, W-2*PAD+12, ROW_H), QColor(r,g,b,90))
        # store code
        p.setPen(QColor(T1)); p.setFont(fRowBold)
        p.drawText(QRectF(col_x[0], ry+6, cols[0][1], ROW_H-6), Qt.AlignmentFlag.AlignLeft, s.code)
        # total/up/fail/pend
        p.setFont(fMono)
        for ci, val in zip(range(1,5), [t, u, f, pd]):
            vcolor = T2 if ci==1 else (GRN if ci==2 else (RED if ci==3 else AMB))
            p.setPen(QColor(vcolor))
            p.drawText(QRectF(col_x[ci], ry+6, cols[ci][1], ROW_H-6), Qt.AlignmentFlag.AlignLeft, str(val))
        # progress bar
        bar_x, bar_w = col_x[5], cols[5][1]-20
        track = QRectF(bar_x, ry+10, bar_w, 10)
        p.setPen(Qt.PenStyle.NoPen); p.setBrush(QColor(SEP))
        p.drawRoundedRect(track, 5, 5)
        fill_w = max(3, bar_w * max(0.0, min(1.0, row_pct)))
        p.setBrush(QColor(color))
        p.drawRoundedRect(QRectF(bar_x, ry+10, fill_w, 10), 5, 5)
        # percent
        p.setPen(QColor(T1)); p.setFont(fRow)
        p.drawText(QRectF(col_x[6], ry+6, cols[6][1], ROW_H-6), Qt.AlignmentFlag.AlignLeft, f"{row_pct*100:.0f}%")

    # -- Footer --
    fy = ty + n*ROW_H + 14
    p.setPen(QPen(QColor(SEP), 1))
    p.drawLine(QPointF(PAD, fy), QPointF(W-PAD, fy))
    p.setPen(QColor(T3)); p.setFont(f2)
    p.drawText(QRectF(PAD, fy+8, W-2*PAD, 20), Qt.AlignmentFlag.AlignLeft, "Generated by Upgrade Command Centre")

    p.end()
    return pix


class SharePanel(QWidget):
    def __init__(self, data: TrackerData, parent=None):
        super().__init__(parent)
        self.data = data
        self.setStyleSheet("background:transparent;")
        self._build()

    def _build(self):
        root = QVBoxLayout(self); root.setContentsMargins(24,20,24,20); root.setSpacing(16)
        eyebrow = make_label("SHARE WITH THE TEAM", T3, FS_SMALL, bold=True); eyebrow.setStyleSheet(eyebrow.styleSheet()+"letter-spacing:2px;")
        root.addWidget(eyebrow)

        row = QHBoxLayout(); row.setSpacing(12)

        copy_card = QWidget(); copy_card.setObjectName("panel")
        copy_card.setStyleSheet(f"QWidget#panel{{{glass_panel_qss(radius=14, border=CY, base_alpha=140, border_alpha=40)}}}")
        cl = QVBoxLayout(copy_card); cl.setContentsMargins(18,16,18,16); cl.setSpacing(10)
        cl.addWidget(make_label("TEXT SUMMARY", T2, 10, bold=True))
        cl.addWidget(make_label("A clean plain-text progress report, ready to paste into Teams, Slack, or an email.", T3, FS_SMALL))
        copy_btn = Btn("📋 Copy summary to clipboard", CY, h=36, filled=True)
        copy_btn.clicked.connect(self._copy_summary)
        cl.addWidget(copy_btn)
        self._preview = QTextEdit(); self._preview.setReadOnly(True)
        self._preview.setStyleSheet(f"QTextEdit{{background:{rgba(BG0,150)};border:1px solid {SEP};border-radius:8px;color:{T2};font-size:11px;font-family:{MONO};padding:8px;}}")
        self._preview.setMinimumHeight(280)
        cl.addWidget(self._preview, 1)
        row.addWidget(copy_card, 1)

        snap_card = QWidget(); snap_card.setObjectName("panel")
        snap_card.setStyleSheet(f"QWidget#panel{{{glass_panel_qss(radius=14, border=VIO, base_alpha=140, border_alpha=40)}}}")
        sl = QVBoxLayout(snap_card); sl.setContentsMargins(18,16,18,16); sl.setSpacing(10)
        sl.addWidget(make_label("PROGRESS REPORT IMAGE", T2, 10, bold=True))
        sl.addWidget(make_label("A standalone summary image with every store listed — sized to fit all of them, built to be read on its own rather than a screenshot of the app.", T3, FS_SMALL))
        snap_btn = Btn("🖼  Save progress report…", VIO, h=36, filled=True)
        snap_btn.clicked.connect(self._save_snapshot)
        sl.addWidget(snap_btn)
        sl.addWidget(make_label("EXCEL EXPORT", T2, 10, bold=True))
        sl.addWidget(make_label("Regenerate a workbook in the original SUMMARY + per-store format, if the team still wants Excel.", T3, FS_SMALL))
        xl_btn = Btn("📊  Export to Excel…", AMB, h=36, filled=True)
        xl_btn.clicked.connect(self._export_xlsx)
        sl.addWidget(xl_btn)
        sl.addStretch()
        self._status_lbl = make_label("", GRN, 9); self._status_lbl.setWordWrap(True)
        sl.addWidget(self._status_lbl)
        row.addWidget(snap_card, 1)

        root.addLayout(row, 1)
        self.refresh_preview()

    def refresh_preview(self):
        self._preview.setPlainText(self._build_summary_text())

    def _build_summary_text(self):
        tot, up, fa, pe = self.data.totals()
        pct = (up/tot*100) if tot else 0
        lines = []
        lines.append(f"POS UPGRADE PROGRESS — {dt.now().strftime('%d %b %Y %H:%M')}")
        lines.append("=" * 46)
        lines.append(f"Overall: {up}/{tot} tills upgraded ({pct:.0f}%) — {fa} failed, {pe} pending")
        lines.append("")
        lines.append(f"{'STORE':<8}{'TOTAL':>7}{'UP':>6}{'FAIL':>6}{'PEND':>6}   STATUS")
        lines.append("-" * 46)
        for s in sorted(self.data.stores.values(), key=lambda s: s.code):
            u, f, p = s.counts()
            t = s.total()
            status = "✓ complete" if (f==0 and p==0) else (f"⚠ {f} failed" if f else f"… {p} pending")
            lines.append(f"{s.code:<8}{t:>7}{u:>6}{f:>6}{p:>6}   {status}")
            if s.notes:
                lines.append(f"         note: {s.notes}")
        lines.append("")
        lines.append("Generated by Upgrade Command Centre")
        return "\n".join(lines)

    def _copy_summary(self):
        text = self._build_summary_text()
        QApplication.clipboard().setText(text)
        self._status_lbl.setStyleSheet(f"color:{GRN};font-size:11px;background:transparent;border:none;")
        self._status_lbl.setText("Copied to clipboard.")
        self.refresh_preview()

    def _save_snapshot(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save progress report", "upgrade_progress.png", "PNG image (*.png)")
        if not path: return
        pix = render_report_image(self.data)
        if pix.save(path):
            self._status_lbl.setStyleSheet(f"color:{GRN};font-size:11px;background:transparent;border:none;")
            self._status_lbl.setText(f"Saved to {path}  ({pix.width()}×{pix.height()}, all {len(self.data.stores)} stores)")
        else:
            self._status_lbl.setStyleSheet(f"color:{RED};font-size:11px;background:transparent;border:none;")
            self._status_lbl.setText("Couldn't save the image.")

    def _export_xlsx(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export to Excel", "UPGRADES_REPORT.xlsx", "Excel files (*.xlsx)")
        if not path: return
        try:
            self._write_xlsx(path)
            self._status_lbl.setStyleSheet(f"color:{GRN};font-size:11px;background:transparent;border:none;")
            self._status_lbl.setText(f"Exported to {path}")
        except Exception as e:
            self._status_lbl.setStyleSheet(f"color:{RED};font-size:11px;background:transparent;border:none;")
            self._status_lbl.setText(f"Export failed: {e}")

    def _safe_sheet_name(self, code, used):
        """Excel sheet titles can't contain  / \\ ? * [ ] :  , can't exceed 31
        chars, and must be unique. Sanitize and de-dupe so export never crashes
        on a real store code."""
        name = code or "STORE"
        for ch in "/\\?*[]:":
            name = name.replace(ch, "-")
        name = name.strip() or "STORE"
        name = name[:31]
        # ensure uniqueness (two codes can collide after truncation/sanitizing)
        if name in used:
            base = name[:28]
            n = 2
            while f"{base}_{n}" in used and n < 999:
                n += 1
            name = f"{base}_{n}"
        used.add(name)
        return name

    def _write_xlsx(self, path):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "SUMMARY"
        headers = ["STORE","TOTAL TILLS","FAILURE TOTAL","UPGRADED","PENDING","FAILED %","UPGRADE %","NOTES"]
        ws.append(headers)
        for c in ws[1]: c.font = Font(bold=True, name="Arial")
        for s in sorted(self.data.stores.values(), key=lambda s: s.code):
            u, f, p = s.counts(); t = s.total()
            ws.append([s.code, t, f, u, p, (f/t if t else 0), (u/t if t else 0), s.notes or ""])
        for row in ws.iter_rows(min_row=2):
            for cell in row: cell.font = Font(name="Arial")
            row[5].number_format = "0.0%"; row[6].number_format = "0.0%"
        for i, w in enumerate([10,11,13,10,9,9,10,50], start=1):
            ws.column_dimensions[chr(64+i)].width = w

        used_names = {"SUMMARY"}
        for s in sorted(self.data.stores.values(), key=lambda s: s.code):
            sh = wb.create_sheet(self._safe_sheet_name(s.code, used_names))
            sh.append(["TILL NUMBER","LAST STATUS","CACHE STATUS","CACHE DETAIL","DATE UPDATED",
                       "POS VERSION","TILL IP","OS_VER","UPGRADE STATUS","POST/PRE CUTOVER FAILURE"])
            for c in sh[1]: c.font = Font(bold=True, name="Arial")
            for t in sorted(s.tills, key=lambda x: x["till"]):
                sh.append([t["till"], t.get("last_status",""), t.get("cache_status",""), t.get("cache_detail",""),
                           t.get("date_updated",""), t.get("pos_version",""), t.get("till_ip",""), t.get("os_ver",""),
                           t.get("upgrade_status",""), t.get("failure_type") or ""])
            for row in sh.iter_rows(min_row=2):
                for cell in row: cell.font = Font(name="Arial")
            for i, w in enumerate([13,13,12,12,12,11,15,9,14,20], start=1):
                sh.column_dimensions[chr(64+i)].width = w
        wb.save(path)


# ── Main window ────────────────────────────────────────────────────────────────
class UpgradeTrackerWindow(QDialog):
    """The full-screen pop-out window. Instantiate once and call show_fullscreen()
    — designed to be launched from a Command Centre Pro HoverCard button."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlag(Qt.WindowType.Window)
        self.setWindowFlag(Qt.WindowType.FramelessWindowHint)
        try:
            from src.widgets.app_icon import apply_window_icon
            apply_window_icon(self)
        except Exception:
            pass
        self.resize(1440, 900)
        self.data = TrackerData()
        self._build()
        if self.data.has_seed_available():
            self._offer_seed_import()

    def show_fullscreen(self):
        self.showMaximized()
        self.show()
        self.activateWindow()

    def _build(self):
        cw = AuroraBackground()
        outer = QVBoxLayout(self); outer.setContentsMargins(0,0,0,0); outer.setSpacing(0)
        outer.addWidget(cw)
        root = QVBoxLayout(cw); root.setContentsMargins(0,0,0,0); root.setSpacing(0)

        root.addWidget(TitleBar(self, self.close))

        nav = QFrame(); nav.setFixedHeight(48)
        nav.setStyleSheet(f"QFrame{{background:{rgba(BG0,150)};border-bottom:1px solid {rgba(SEP,110)};}}")
        nl = QHBoxLayout(nav); nl.setContentsMargins(16,8,16,8); nl.setSpacing(4)
        pill = QFrame(); pill.setFixedHeight(40)
        pill.setStyleSheet(f"QFrame{{{glass_panel_qss(radius=16, base=BG2, border=SEP, base_alpha=140, border_alpha=90)}}}")
        pl = QHBoxLayout(pill); pl.setContentsMargins(3,3,3,3); pl.setSpacing(2)
        self._nav_btns = {}
        for label, key in [("Dashboard","dash"), ("Stores","stores"), ("Import","import"), ("Share","share")]:
            b = QPushButton(label); b.setFixedHeight(34); b.setCheckable(True)
            b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            b.clicked.connect(lambda _, k=key: self._nav(k))
            # Font-weight is deliberately the SAME (700) in every state below —
            # QPushButton's sizeHint()/layout width is computed once and isn't
            # reliably re-measured when only a QSS pseudo-state like :checked
            # changes font-weight, so a lighter unchecked weight next to a
            # bolder checked weight can render wider than the space reserved
            # for it (exactly the clipped "Stores" pill). Only color/background
            # differ between states now, so the reserved width always matches.
            b.setStyleSheet(f"QPushButton{{background:transparent;border:none;color:{T3};padding:0 20px;font-size:12px;font-weight:700;border-radius:12px;}}"
                             f"QPushButton:checked{{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 {CY},stop:1 {VIO});color:{BG0};}}"
                             f"QPushButton:hover:!checked{{color:{T1};background:{rgba(BG4,140)};}}")
            pl.addWidget(b); self._nav_btns[key] = b
        nl.addWidget(pill)
        nl.addStretch()
        self._path_lbl = make_label(data_path(), T3, FS_TINY, mono=True)
        nl.addWidget(self._path_lbl)
        root.addWidget(nav)

        self._stack = QStackedWidget(); self._stack.setStyleSheet("background:transparent;")
        self._dash = DashboardPanel(self.data)
        self._stores = StoresPanel(self.data)
        self._import = ImportPanel(self.data)
        self._import.imported.connect(self._on_data_changed)
        self._share = SharePanel(self.data)
        for w in [self._dash, self._stores, self._import, self._share]:
            self._stack.addWidget(w)
        root.addWidget(self._stack, 1)

        self._nav_btns["dash"].setChecked(True)

    def _nav(self, key):
        idx = {"dash":0, "stores":1, "import":2, "share":3}[key]
        self._stack.setCurrentIndex(idx)
        for k, b in self._nav_btns.items(): b.setChecked(k==key)
        if key == "dash": self._dash.refresh()
        if key == "stores": self._stores.refresh_list()
        if key == "share": self._share.refresh_preview()

    def _on_data_changed(self):
        self._dash.refresh()
        self._stores.refresh_list()

    def _offer_seed_import(self):
        box = QMessageBox(self)
        box.setWindowTitle("Import existing data?")
        box.setText(f"Found existing tracking data ({len(json.load(open(seed_path()))['stores'])} stores) "
                     "from your previous workbook. Import it now?")
        box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if box.exec() == QMessageBox.StandardButton.Yes:
            self.data.import_seed()
            self._on_data_changed()

    def keyPressEvent(self, e):
        if e.key() == Qt.Key.Key_Escape:
            if self.isMaximized(): self.showNormal()
            else: super().keyPressEvent(e)
        else:
            super().keyPressEvent(e)


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setFont(QFont("Segoe UI", 10))
    win = UpgradeTrackerWindow()
    win.show_fullscreen()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
